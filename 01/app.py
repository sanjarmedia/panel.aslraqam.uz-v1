from flask import Flask, request, jsonify, render_template_string, Response
import requests
import time
import threading
import zipfile
import json
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# New API server (migrated from goods.aslbelgisi.uz on 2025-12-01)
BASE_URL = "https://xtrace.aslbelgisi.uz"
CHUNK_SIZE = 1000
INTER_CHUNK_SLEEP = 0.3   # seconds between chunks sent to ASL API
MAX_RETRIES = 3            # retries on 429 Too Many Requests

# Limit concurrent outbound API calls: at most 2 simultaneous requests to ASL
_api_sem = threading.Semaphore(2)


def _api_call(method, url, **kwargs):
    """HTTP request to ASL API with retry on 429 (exponential backoff)."""
    backoff = 2.0
    for attempt in range(MAX_RETRIES + 1):
        with _api_sem:
            try:
                resp = requests.request(method, url, timeout=30, **kwargs)
            except requests.exceptions.RequestException:
                raise
        if resp.status_code == 429 and attempt < MAX_RETRIES:
            time.sleep(backoff)
            backoff *= 2
            continue
        return resp
    return resp

# ── Routes ──────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template_string(HTML)

@app.route("/api/info", methods=["POST"])
def proxy_info():
    body = request.get_json(force=True)
    codes = body.get("codes", [])
    token = body.get("token", "")

    if not token:
        return jsonify({"error": "Токен не указан"}), 400
    if not codes:
        return jsonify([])

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json;charset=UTF-8"
    }

    results = []
    for i in range(0, len(codes), CHUNK_SIZE):
        chunk = codes[i:i + CHUNK_SIZE]
        try:
            resp = _api_call(
                "POST",
                f"{BASE_URL}/public/api/cod/public/codes",
                headers=headers,
                json={"codes": chunk, "addCodeHistory": False},
            )
        except requests.exceptions.RequestException as e:
            return jsonify({"error": str(e)}), 502

        if not resp.ok:
            return jsonify({"error": resp.text, "status": resp.status_code}), resp.status_code
        results.extend(resp.json())

        if i + CHUNK_SIZE < len(codes):
            time.sleep(INTER_CHUNK_SLEEP)

    return jsonify(results)


@app.route("/api/history", methods=["GET"])
def proxy_history():
    code = request.args.get("code", "")
    token = request.args.get("token", "")

    if not token:
        return jsonify({"error": "Токен не указан"}), 400

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json;charset=UTF-8"
    }
    try:
        resp = requests.post(
            f"{BASE_URL}/public/api/cod/public/codes",
            headers=headers,
            json={"codes": [code], "addCodeHistory": True},
            timeout=30
        )
    except requests.exceptions.RequestException as e:
        return jsonify({"error": str(e)}), 502

    if not resp.ok:
        return jsonify({"error": resp.text, "status": resp.status_code}), resp.status_code

    data = resp.json()
    if not data:
        return jsonify([])
    return jsonify(data[0].get("codeHistory", []))


# ── Documents (section 11) ───────────────────────────────────────────────────

def _auth_headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json;charset=UTF-8"}


def _extract_list(data):
    """Extract list of items from any response shape."""
    if isinstance(data, list):
        return data
    for key in ("receipts", "items", "content", "data", "result", "list"):
        if key in data and isinstance(data[key], list):
            return data[key]
    return []


_rcp_cache = {"receipts": [], "synced_at": None, "total": 0, "token": ""}


@app.route("/api/rcp-sync", methods=["POST"])
def rcp_sync():
    body = request.get_json(force=True)
    token = body.get("token", "")
    if not token:
        return jsonify({"error": "no token"}), 400

    all_receipts = []
    cursor = None
    while True:
        params = [("limit", "100")]
        if cursor:
            params.append(("cursor", cursor))
        try:
            resp = requests.get(
                f"{BASE_URL}/api/rcp/receipts",
                headers=_auth_headers(token),
                params=params,
                timeout=30
            )
        except requests.exceptions.RequestException as e:
            if all_receipts:
                break
            return jsonify({"error": str(e)}), 502

        if not resp.ok:
            if all_receipts:
                break
            return jsonify({"error": resp.text}), resp.status_code

        data = resp.json()
        items = _extract_list(data)
        all_receipts.extend(items)
        next_cursor = None
        if isinstance(data, dict):
            next_cursor = data.get("cursor") or data.get("nextCursor")
        if not next_cursor or len(items) < 100:
            break
        cursor = next_cursor

    _rcp_cache["receipts"] = all_receipts
    _rcp_cache["total"] = len(all_receipts)
    _rcp_cache["synced_at"] = time.strftime("%Y-%m-%dT%H:%M:%S")
    _rcp_cache["token"] = token
    return jsonify({"count": len(all_receipts), "synced_at": _rcp_cache["synced_at"]})


@app.route("/api/rcp-cached", methods=["GET"])
def rcp_cached():
    items = list(_rcp_cache["receipts"])
    date_from = request.args.get("dateFrom")
    date_to   = request.args.get("dateTo")
    rcp_type  = request.args.get("type", "").upper()
    ext_id    = request.args.get("externalId", "").strip().lower()

    if date_from:
        items = [r for r in items if (r.get("createdOn") or r.get("createdDate") or "") >= date_from]
    if date_to:
        items = [r for r in items if (r.get("createdOn") or r.get("createdDate") or "") <= date_to + "T23:59:59"]
    if rcp_type:
        items = [r for r in items if (r.get("type") or "").upper() == rcp_type]
    if ext_id:
        items = [r for r in items if ext_id in (r.get("externalId") or "").lower()]

    return jsonify({
        "receipts": items,
        "total": len(items),
        "cached_total": _rcp_cache["total"],
        "synced_at": _rcp_cache["synced_at"]
    })


@app.route("/api/rcp-detail/<receipt_id>", methods=["GET"])
def rcp_detail(receipt_id):
    token = _rcp_cache.get("token", "")
    if not token:
        return jsonify({"error": "Нет токена — выполните синхронизацию заново"}), 400
    try:
        resp = requests.get(
            f"{BASE_URL}/api/rcp/receipts/{receipt_id}",
            headers=_auth_headers(token),
            timeout=30
        )
    except requests.exceptions.RequestException as e:
        return jsonify({"error": str(e)}), 502
    if not resp.ok:
        return jsonify({"error": resp.text}), resp.status_code
    return jsonify(resp.json())


@app.route("/api/docs", methods=["GET"])
def proxy_docs():
    token = request.args.get("token", "")
    if not token:
        return jsonify({"error": "Токен не указан"}), 400

    params = []
    for k in ("dateFrom", "dateTo", "limit", "cursor"):
        if request.args.get(k):
            params.append((k, request.args.get(k)))
    for v in request.args.getlist("status"):
        params.append(("status", v))
    for v in request.args.getlist("types"):
        params.append(("types", v))

    try:
        resp = requests.get(
            f"{BASE_URL}/public/api/v1/doc/storage/docs/search",
            headers=_auth_headers(token),
            params=params,
            timeout=30
        )
    except requests.exceptions.RequestException as e:
        return jsonify({"error": str(e)}), 502

    if not resp.ok:
        return jsonify({"error": resp.text}), resp.status_code
    return jsonify(resp.json())


@app.route("/api/docs/<doc_id>", methods=["GET"])
def proxy_doc_meta(doc_id):
    token = request.args.get("token", "")
    if not token:
        return jsonify({"error": "Токен не указан"}), 400
    try:
        resp = requests.get(
            f"{BASE_URL}/public/api/v1/doc/storage/docs/{doc_id}",
            headers=_auth_headers(token), timeout=30
        )
    except requests.exceptions.RequestException as e:
        return jsonify({"error": str(e)}), 502
    if not resp.ok:
        return jsonify({"error": resp.text}), resp.status_code
    return jsonify(resp.json())


@app.route("/api/docs/<doc_id>/errors", methods=["GET"])
def proxy_doc_errors(doc_id):
    token = request.args.get("token", "")
    if not token:
        return jsonify({"error": "Токен не указан"}), 400
    params = {
        "limit": request.args.get("limit", 100),
    }
    if request.args.get("lastIndex"):
        params["lastIndex"] = request.args.get("lastIndex")
    try:
        resp = requests.get(
            f"{BASE_URL}/public/api/v1/doc/storage/errors/{doc_id}",
            headers=_auth_headers(token), params=params, timeout=30
        )
    except requests.exceptions.RequestException as e:
        return jsonify({"error": str(e)}), 502
    if not resp.ok:
        return jsonify({"error": resp.text}), resp.status_code
    return jsonify(resp.json())


@app.route("/api/docs/<doc_id>/codes", methods=["GET"])
def proxy_doc_codes(doc_id):
    token = request.args.get("token", "")
    if not token:
        return jsonify({"error": "Токен не указан"}), 400
    params = {"limit": request.args.get("limit", 200)}
    if request.args.get("lastIndex"):
        params["lastIndex"] = request.args.get("lastIndex")
    try:
        resp = requests.get(
            f"{BASE_URL}/public/api/v1/doc/storage/docs/{doc_id}/codes",
            headers=_auth_headers(token), params=params, timeout=30
        )
    except requests.exceptions.RequestException as e:
        return jsonify({"error": str(e)}), 502
    if not resp.ok:
        return jsonify({"error": resp.text}), resp.status_code
    return jsonify(resp.json())


# ── Bulk export (section 8) ──────────────────────────────────────────────────

@app.route("/api/export", methods=["POST"])
def proxy_export_create():
    body = request.get_json(force=True)
    token = body.pop("token", "")
    if not token:
        return jsonify({"error": "Токен не указан"}), 400
    try:
        resp = requests.post(
            f"{BASE_URL}/public/api/cod/exports",
            headers=_auth_headers(token),
            json=body,
            timeout=30
        )
    except requests.exceptions.RequestException as e:
        return jsonify({"error": str(e)}), 502
    if not resp.ok:
        return jsonify({"error": resp.text}), resp.status_code
    return jsonify(resp.json())


@app.route("/api/export/<job_id>/status", methods=["GET"])
def proxy_export_status(job_id):
    token = request.args.get("token", "")
    if not token:
        return jsonify({"error": "Токен не указан"}), 400
    try:
        resp = requests.get(
            f"{BASE_URL}/public/api/cod/exports/{job_id}/status",
            headers=_auth_headers(token), timeout=30
        )
    except requests.exceptions.RequestException as e:
        return jsonify({"error": str(e)}), 502
    if not resp.ok:
        return jsonify({"error": resp.text}), resp.status_code
    return jsonify(resp.json())


_STATUS_RU = {
    "INTRODUCED": "В обороте", "WITHDRAWN": "Выведен из оборота",
    "APPLIED": "Нанесён", "RECEIVED": "Получен", "WRITTEN_OFF": "Списан",
    "EMITTED": "Эмитирован", "RETIRED": "Аннулирован",
}
_PKG_RU = {
    "UNIT": "Единица", "GROUP": "Группа",
    "BOX_LV_1": "Короб", "BOX_LV_2": "Паллет",
}
_RELEASE_RU = {
    "IMPORT": "Импорт", "PRODUCTION": "Производство",
    "PRIMARY": "Первичный выпуск", "COMMISSION": "Комиссионный",
    "REMAINS": "Остатки", "CROSSBORDER": "Трансграничный",
}
_COUNTRY_RU = {
    "cn": "Китай", "ru": "Россия", "kz": "Казахстан", "uz": "Узбекистан",
    "de": "Германия", "tr": "Турция", "us": "США", "kr": "Корея",
    "it": "Италия", "fr": "Франция", "gb": "Великобритания",
}


def _g(d, *keys, default=""):
    """Safe nested get."""
    for k in keys:
        if not isinstance(d, dict):
            return default
        d = d.get(k, None)
        if d is None:
            return default
    return d if d is not None else default


def _date(val):
    return (val or "")[:10]


# Columns: (header, getter(record))
_COLUMNS = [
    ("Код маркировки",          lambda r: _g(r, "codeData", "code")),
    ("Статус",                  lambda r: _STATUS_RU.get(_g(r, "codeData", "status"), _g(r, "codeData", "status"))),
    ("Тип упаковки",            lambda r: _PKG_RU.get(_g(r, "packageData", "packageType"), _g(r, "packageData", "packageType"))),
    ("GTIN",                    lambda r: _g(r, "productData", "gtin")),
    ("Дата эмиссии",            lambda r: _date(_g(r, "markingData", "emissionDate"))),
    ("Дата производства",       lambda r: _date(_g(r, "productData", "productionDate"))),
    ("Срок годности",           lambda r: _date(_g(r, "productData", "expirationDate"))),
    ("Страна производства",     lambda r: _COUNTRY_RU.get(_g(r, "productData", "manufacturerCountry"),
                                                          _g(r, "productData", "manufacturerCountry"))),
    ("Способ ввода в оборот",   lambda r: _RELEASE_RU.get(_g(r, "turnoverData", "originalReleaseMethod"),
                                                          _g(r, "turnoverData", "originalReleaseMethod"))),
    ("ИНН эмитента",            lambda r: _g(r, "markingData", "issuerInfo", "issuerTin")),
    ("Эмитент",                 lambda r: _g(r, "markingData", "issuerInfo", "issuerName", "ru")),
    ("ИНН владельца",           lambda r: _g(r, "turnoverData", "ownerInfo", "ownerTin")),
    ("Владелец",                lambda r: _g(r, "turnoverData", "ownerInfo", "ownerName", "ru")),
    ("Дата нанесения",          lambda r: _date(_g(r, "markingData", "utilisationDate"))),
]


def _records_from_zip(zip_bytes):
    """Extract records list from ASL export ZIP (export.json → data['results'])."""
    records = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for name in zf.namelist():
            raw = zf.read(name)
            try:
                data = json.loads(raw)
            except Exception:
                continue
            if isinstance(data, list):
                records.extend(data)
            elif isinstance(data, dict):
                # Real structure: {"results": [...]}
                for key in ("results", "codes", "items", "data"):
                    if key in data and isinstance(data[key], list):
                        records.extend(data[key])
                        break
                else:
                    records.append(data)
    return records


def _build_excel(records, job_id):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Коды маркировки"

    # ── Header row ──────────────────────────────────────────────────────────
    hdr_fill = PatternFill("solid", fgColor="0B6E4F")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_border = Border(
        bottom=Side(style="medium", color="C4F542"),
    )
    for col_idx, (title, _) in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = hdr_border
    ws.row_dimensions[1].height = 30

    # ── Data rows ────────────────────────────────────────────────────────────
    alt_fill = PatternFill("solid", fgColor="1A2E28")
    for row_idx, record in enumerate(records, 2):
        for col_idx, (_, getter) in enumerate(_COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=getter(record))
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # ── Column widths ────────────────────────────────────────────────────────
    widths = [52, 18, 14, 16, 13, 14, 13, 18, 22, 14, 34, 14, 34, 13]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Autofilter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    # ── Freeze top row ───────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Сводка")
    from collections import Counter
    status_cnt = Counter(_STATUS_RU.get(r.get("status", ""), r.get("status", "")) for r in records)
    ws2.cell(1, 1, "Статус").font = Font(bold=True)
    ws2.cell(1, 2, "Количество").font = Font(bold=True)
    for i, (st, cnt) in enumerate(sorted(status_cnt.items()), 2):
        ws2.cell(i, 1, st)
        ws2.cell(i, 2, cnt)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@app.route("/api/export/<job_id>/result", methods=["GET"])
def proxy_export_result(job_id):
    token = request.args.get("token", "")
    if not token:
        return jsonify({"error": "Токен не указан"}), 400
    try:
        resp = requests.get(
            f"{BASE_URL}/public/api/cod/exports/{job_id}/result",
            headers={"Authorization": f"Bearer {token}"},
            timeout=120
        )
    except requests.exceptions.RequestException as e:
        return jsonify({"error": str(e)}), 502
    if not resp.ok:
        return jsonify({"error": resp.text}), resp.status_code

    try:
        records = _records_from_zip(resp.content)
        xlsx_bytes = _build_excel(records, job_id)
    except Exception as e:
        return jsonify({"error": f"Ошибка конвертации: {e}"}), 500

    return Response(
        xlsx_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=export-{job_id}.xlsx"}
    )


@app.route("/api/export/<job_id>/debug", methods=["GET"])
def debug_export_zip(job_id):
    token = request.args.get("token", "")
    if not token:
        return jsonify({"error": "Токен не указан"}), 400
    try:
        resp = requests.get(
            f"{BASE_URL}/public/api/cod/exports/{job_id}/result",
            headers={"Authorization": f"Bearer {token}"},
            timeout=120
        )
    except requests.exceptions.RequestException as e:
        return jsonify({"error": str(e)}), 502
    if not resp.ok:
        return jsonify({"error": resp.text}), resp.status_code
    try:
        with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
            names = zf.namelist()
            previews = {}
            for name in names:
                raw = zf.read(name)
                try:
                    parsed = json.loads(raw)
                    if isinstance(parsed, list):
                        previews[name] = {"type": "list", "len": len(parsed), "first": parsed[0] if parsed else None}
                    elif isinstance(parsed, dict):
                        previews[name] = {"type": "dict", "keys": list(parsed.keys()),
                                          "sample": {k: (v[:1] if isinstance(v, list) else v)
                                                     for k, v in parsed.items()}}
                    else:
                        previews[name] = {"type": type(parsed).__name__, "value": str(parsed)[:200]}
                except Exception as ex:
                    previews[name] = {"type": "raw", "error": str(ex), "preview": raw[:200].decode("utf-8", errors="replace")}
        return jsonify({"files": names, "previews": previews})
    except Exception as e:
        return jsonify({"error": str(e), "content_type": resp.headers.get("Content-Type"),
                        "content_preview": resp.content[:500].decode("utf-8", errors="replace")}), 500


_CODE_STATE_RU = {
    "SUCCESS": "Успешно", "ERROR": "Ошибка", "WARNING": "Предупреждение",
}

_STATUS_DOC_RU = {
    "SUCCESS": "Успешно", "PARTIALLY_PROCESSED": "Частично обработан",
    "ERROR": "Ошибка", "IN_PROCESS": "В обработке",
    "IN_PROCESSING": "Обрабатывается", "VALIDATING": "Валидация",
    "CREATED": "Создан", "EXPIRED": "Просрочен",
}
_DOC_TYPE_RU = {
    "ORIGINAL_INVOICE": "ЭСФ (стандартная)", "ADDITIONAL_INVOICE": "ЭСФ (дополнительная)",
    "CORRECTED_INVOICE": "ЭСФ (исправленная)", "CANCELLED_INVOICE": "ЭСФ (аннулированная)",
    "SALES_RECEIPT": "Чек ККТ (продажа)", "REFUND_RECEIPT": "Чек ККТ (возврат)",
    "ORDER": "Заказ КМ",
    "ORIGINAL_CUSTOMS_DECLARATION": "Таможенная декларация",
    "CUSTOMS_CODE_WITHDRAWAL": "Вывод через таможню",
    "CUSTOMS_CODE_REGISTRATION": "Регистрация АИК",
    "UTILISATION": "Нанесение КМ", "AGGREGATION": "Агрегация",
    "TRANSFER_REQUEST": "Передача КМ", "WRITE_OFF_NOTICE": "Списание",
    "WITHDRAWAL": "Вывод из оборота", "VALIDATION": "Валидация печати",
    "TRANSPORT_CODE_DISAGGREGATION": "Расформирование",
}


@app.route("/api/docs-list-excel", methods=["POST"])
def docs_list_excel():
    body = request.get_json(force=True)
    docs = body.get("docs", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Документы"

    hdr_fill = PatternFill("solid", fgColor="0B6E4F")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["#", "Doc ID", "Тип документа", "Статус", "Дата создания", "Внешний ID"]
    col_widths = [5, 52, 26, 22, 20, 36]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 26

    alt_fill = PatternFill("solid", fgColor="1A2E28")
    for ri, doc in enumerate(docs, 2):
        create_date = (doc.get("createDate") or "")[:19].replace("T", " ")
        row_data = [
            ri - 1,
            doc.get("documentId", ""),
            _DOC_TYPE_RU.get(doc.get("type", ""), doc.get("type", "")),
            _STATUS_DOC_RU.get(doc.get("status", ""), doc.get("status", "")),
            create_date,
            doc.get("originalDocId", "") or "",
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                cell.fill = alt_fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dokumenty.xlsx"}
    )


@app.route("/api/rcp-bulk-codes-excel", methods=["POST"])
def rcp_bulk_codes_excel():
    body = request.get_json(force=True)
    receipts = body.get("receipts", [])  # [{receiptId, externalId, type, codes:[{code,state,gtin}]}]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Коды КМ — все чеки"

    _state_ru = {
        "INTRODUCED": "В обороте", "WITHDRAWN": "Выведен из оборота",
        "APPLIED": "Нанесён", "RECEIVED": "Получен", "WRITTEN_OFF": "Списан",
        "CHANGED": "Изменено", "RETIRED": "Аннулирован", "EMITTED": "Эмитирован",
        "SUCCESS": "Успешно", "ERROR": "Ошибка", "WARNING": "Предупреждение",
    }
    _type_ru = {"SALE": "Продажа", "RETURN": "Возврат", "SALES": "Продажа", "REFUND": "Возврат"}

    hdr_fill = PatternFill("solid", fgColor="0B6E4F")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    headers  = ["#", "ID чека", "Внешний ID (ФН)", "Тип чека", "Код маркировки", "GTIN", "Состояние КМ"]
    col_widths = [6, 42, 32, 12, 60, 18, 20]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 26

    alt_fill = PatternFill("solid", fgColor="1A2E28")
    ri = 2
    for rcp in receipts:
        rid    = rcp.get("receiptId", "")
        extid  = rcp.get("externalId", "")
        rtype  = _type_ru.get((rcp.get("type") or "").upper(), rcp.get("type", ""))
        for item in rcp.get("codes", []):
            state_raw = item.get("state", "")
            row_data = [
                ri - 1, rid, extid, rtype,
                item.get("code", ""),
                item.get("gtin", ""),
                _state_ru.get(state_raw, state_raw),
            ]
            for ci2, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci2, value=val)
                cell.alignment = Alignment(vertical="center")
                if ri % 2 == 0:
                    cell.fill = alt_fill
            ri += 1

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Сводка по чекам")
    ws2.cell(1, 1, "ID чека").font = Font(bold=True)
    ws2.cell(1, 2, "Внешний ID").font = Font(bold=True)
    ws2.cell(1, 3, "Тип").font = Font(bold=True)
    ws2.cell(1, 4, "Всего КМ").font = Font(bold=True)
    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 32
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 12
    for si, rcp in enumerate(receipts, 2):
        ws2.cell(si, 1, rcp.get("receiptId", ""))
        ws2.cell(si, 2, rcp.get("externalId", ""))
        ws2.cell(si, 3, _type_ru.get((rcp.get("type") or "").upper(), rcp.get("type", "")))
        ws2.cell(si, 4, len(rcp.get("codes", [])))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=kody-kkt.xlsx"}
    )


@app.route("/api/receipts", methods=["GET"])
def proxy_receipts():
    token = request.args.get("token", "")
    if not token:
        return jsonify({"error": "Токен не указан"}), 400
    params = []
    for k in ("createdOnFrom", "createdOnTo", "limit", "cursor", "type", "status", "sellerId"):
        if request.args.get(k):
            params.append((k, request.args.get(k)))
    try:
        resp = requests.get(
            f"{BASE_URL}/api/rcp/receipts",
            headers=_auth_headers(token),
            params=params,
            timeout=30
        )
    except requests.exceptions.RequestException as e:
        return jsonify({"error": str(e)}), 502
    if not resp.ok:
        return jsonify({"error": resp.text}), resp.status_code
    return jsonify(resp.json())


@app.route("/api/receipts/<receipt_id>", methods=["GET"])
def proxy_receipt_detail(receipt_id):
    token = request.args.get("token", "")
    if not token:
        return jsonify({"error": "Токен не указан"}), 400
    try:
        resp = requests.get(
            f"{BASE_URL}/api/rcp/receipts/{receipt_id}",
            headers=_auth_headers(token),
            timeout=30
        )
    except requests.exceptions.RequestException as e:
        return jsonify({"error": str(e)}), 502
    if not resp.ok:
        return jsonify({"error": resp.text}), resp.status_code
    return jsonify(resp.json())


@app.route("/api/receipts-excel", methods=["POST"])
def receipts_excel():
    body = request.get_json(force=True)
    receipts = body.get("receipts", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Чеки ККТ"

    hdr_fill = PatternFill("solid", fgColor="0B6E4F")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["#", "ID чека", "Внешний ID", "Тип", "Статус", "Дата регистрации", "Дата чека", "ИНН продавца", "Продавец"]
    col_widths = [5, 42, 32, 14, 22, 20, 20, 14, 36]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 26

    _type_ru = {"SALE": "Продажа", "RETURN": "Возврат", "SALES": "Продажа", "REFUND": "Возврат"}
    _status_ru = {"SUCCESS": "Обработан успешно", "ERROR": "Ошибка", "IN_PROCESS": "В обработке",
                  "PARTIALLY_PROCESSED": "Частично", "CREATED": "Создан"}
    alt_fill = PatternFill("solid", fgColor="1A2E28")
    for ri, r in enumerate(receipts, 2):
        seller = r.get("seller") or r.get("sellerInfo") or {}
        seller_tin = seller.get("tin") or seller.get("sellerTin") or r.get("sellerTin", "")
        seller_name = seller.get("name") or seller.get("sellerName") or r.get("sellerName", "")
        if isinstance(seller_name, dict):
            seller_name = seller_name.get("ru") or seller_name.get("uz") or ""
        rec_type = _type_ru.get(r.get("type", ""), r.get("type", ""))
        rec_status = _status_ru.get(r.get("status", ""), r.get("status", ""))
        created = (r.get("createdOn") or r.get("createdDate") or "")[:19].replace("T", " ")
        receipt_dt = (r.get("receiptDate") or r.get("dateTime") or r.get("checkDate") or "")[:19].replace("T", " ")
        row_data = [
            ri - 1,
            r.get("id") or r.get("receiptId") or "",
            r.get("externalId") or r.get("originalDocId") or "",
            rec_type, rec_status, created, receipt_dt,
            seller_tin, seller_name,
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                cell.fill = alt_fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cheki-kkt.xlsx"}
    )


@app.route("/api/doc-codes-excel", methods=["POST"])
def doc_codes_excel():
    body = request.get_json(force=True)
    codes = body.get("codes", [])
    filename = body.get("filename", "doc-codes.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Коды документа"

    hdr_fill = PatternFill("solid", fgColor="0B6E4F")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["#", "Код маркировки", "Состояние", "Причина ошибки"]
    col_widths = [6, 56, 16, 40]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 26

    alt_fill = PatternFill("solid", fgColor="1A2E28")
    for ri, item in enumerate(codes, 2):
        state_raw = item.get("state", "")
        state_ru = _CODE_STATE_RU.get(state_raw, state_raw)
        row_data = [ri - 1, item.get("code", ""), state_ru, item.get("errorCode", "") or ""]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                cell.fill = alt_fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.route("/api/docs-bulk-codes-excel", methods=["POST"])
def docs_bulk_codes_excel():
    body = request.get_json(force=True)
    docs = body.get("docs", [])   # [{facturaNo, facturaId, docId, type, codes:[{code,state,errorCode}]}]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Коды КМ — все ЭСФ"

    hdr_fill = PatternFill("solid", fgColor="0B6E4F")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    headers  = ["#", "Номер ЭСФ", "FacturaId", "Тип документа", "Код маркировки", "Состояние", "Причина ошибки"]
    col_widths = [6, 18, 30, 24, 58, 16, 34]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 26

    alt_fill = PatternFill("solid", fgColor="1A2E28")
    ri = 2
    for doc in docs:
        factura_no = doc.get("facturaNo", "")
        factura_id = doc.get("facturaId", "")
        doc_type   = _DOC_TYPE_RU.get(doc.get("type", ""), doc.get("type", ""))
        for item in doc.get("codes", []):
            state_raw = item.get("state", "")
            row_data = [
                ri - 1,
                factura_no,
                factura_id,
                doc_type,
                item.get("code", ""),
                _CODE_STATE_RU.get(state_raw, state_raw),
                item.get("errorCode", "") or "",
            ]
            for ci2, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci2, value=val)
                cell.alignment = Alignment(vertical="center")
                if ri % 2 == 0:
                    cell.fill = alt_fill
            ri += 1

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Summary sheet: per-document code counts
    ws2 = wb.create_sheet("Сводка по ЭСФ")
    ws2.cell(1, 1, "Номер ЭСФ").font = Font(bold=True)
    ws2.cell(1, 2, "FacturaId").font = Font(bold=True)
    ws2.cell(1, 3, "Всего КМ").font = Font(bold=True)
    ws2.cell(1, 4, "Успешно").font = Font(bold=True)
    ws2.cell(1, 5, "Ошибка").font = Font(bold=True)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 12
    for si, doc in enumerate(docs, 2):
        codes = doc.get("codes", [])
        ok  = sum(1 for c in codes if c.get("state") == "SUCCESS")
        err = sum(1 for c in codes if c.get("state") in ("ERROR", "WARNING"))
        ws2.cell(si, 1, doc.get("facturaNo", ""))
        ws2.cell(si, 2, doc.get("facturaId", ""))
        ws2.cell(si, 3, len(codes))
        ws2.cell(si, 4, ok)
        ws2.cell(si, 5, err)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=vse-kody-esf.xlsx"}
    )


# ── Counterparty (section 12) ────────────────────────────────────────────────

@app.route("/api/party/<tin>", methods=["GET"])
def proxy_party(tin):
    token = request.args.get("token", "")
    if not token:
        return jsonify({"error": "Токен не указан"}), 400
    try:
        resp = requests.get(
            f"{BASE_URL}/public/api/v1/party/parties/{tin}/status",
            headers=_auth_headers(token), timeout=30
        )
    except requests.exceptions.RequestException as e:
        return jsonify({"error": str(e)}), 502
    if not resp.ok:
        return jsonify({"error": resp.text}), resp.status_code
    return jsonify(resp.json())


# ── HTML Template ────────────────────────────────────────────────────────────
# New API response format (xtrace.aslbelgisi.uz /public/api/cod/public/codes):
#   Array of: { code, status, extendedStatus, packageType, gtin,
#               expirationDate, emissionDate,
#               issuerShortInfo: { issuerTin, issuerName:{ru,uz,en} } }
# History uses same endpoint with addCodeHistory:true → item.codeHistory[]
#   codeHistory item: { eventDate, eventType, eventSourceId,
#                       senderTin, receiverTin, eventChangedCodeStatus }

HTML = r"""<!DOCTYPE html>
<html lang="ru" class="dark">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>ASL RAQAM — Проверка кодов маркировки</title>
  <script src="https://cdn.tailwindcss.com"></script>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link href="https://fonts.googleapis.com/css2?family=Unbounded:wght@400;700&family=Manrope:wght@400;500;600&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
  <style>
    *{box-sizing:border-box}
    body{font-family:'Manrope',sans-serif;background:#0E1712;color:#e5e7eb;margin:0}
    .dm-grid-bg{background-image:linear-gradient(rgba(196,245,66,.04) 1px,transparent 1px),linear-gradient(90deg,rgba(196,245,66,.04) 1px,transparent 1px);background-size:32px 32px}
    .header-bar{border-bottom:1px solid rgba(11,110,79,.25);padding-bottom:20px;margin-bottom:24px;position:relative}
    .header-bar::after{content:'';position:absolute;bottom:-1px;left:0;width:180px;height:2px;background:linear-gradient(90deg,#0B6E4F,#C4F542,transparent)}
    .font-display{font-family:'Unbounded',sans-serif}
    .mono{font-family:'JetBrains Mono',monospace}
    ::-webkit-scrollbar{width:5px;height:5px}
    ::-webkit-scrollbar-track{background:#111}
    ::-webkit-scrollbar-thumb{background:#0B6E4F;border-radius:3px}
    @keyframes fadeUp{from{opacity:0;transform:translateY(5px)}to{opacity:1;transform:translateY(0)}}
    .row-enter{animation:fadeUp .18s ease forwards}
    @keyframes pulse-bar{0%,100%{opacity:1}50%{opacity:.45}}
    .bar-pulse{animation:pulse-bar 1.4s ease-in-out infinite}
    .tab-active{background:#0B6E4F!important;color:#C4F542!important;font-weight:600}
    .subtab-active{background:#0B6E4F!important;color:#C4F542!important;font-weight:600}
    .filter-pill{padding:5px 12px;border-radius:20px;border:1px solid #1f2937;color:#6b7280;
                 font-size:12px;cursor:pointer;transition:all .15s;white-space:nowrap}
    .filter-pill:hover{border-color:#374151;color:#9ca3af}
    .filter-pill.active{border-color:#0B6E4F;background:rgba(11,110,79,.2);color:#C4F542}
    .result-row{transition:background .12s}
    .result-row:hover{background:rgba(11,110,79,.07)}
    .result-row.expanded-open{background:rgba(11,110,79,.05)}
    .expand-row{background:#0d0d0d}
    .status-badge{display:inline-flex;align-items:center;padding:2px 8px;border-radius:6px;
                  font-size:11px;font-weight:500;border:1px solid;white-space:nowrap}
    .copy-cell{cursor:copy}
    .copy-cell:active{opacity:.6}
    @keyframes copy-flash{0%{background:rgba(196,245,66,.25)}100%{background:transparent}}
    .flash{animation:copy-flash .5s ease}
    .card{background:#0A0F0C;border:1px solid rgba(11,110,79,.18);border-radius:12px}
    input[type=checkbox]{accent-color:#0B6E4F;width:14px;height:14px}
    .timeline-dot{width:10px;height:10px;border-radius:50%;border:2px solid #0B6E4F;
                  background:#080808;flex-shrink:0;margin-top:4px}
    .timeline-line{width:1px;background:#1f2937;flex:1;margin:4px auto}
    textarea:focus,input:focus{outline:none;border-color:#0B6E4F!important}
    .btn-primary{background:#0B6E4F;color:#C4F542;font-weight:600;padding:8px 20px;
                 border-radius:8px;font-size:13px;transition:all .15s;cursor:pointer}
    .btn-primary:hover{background:#0d7d5a}
    .btn-primary:disabled{opacity:.4;cursor:not-allowed}
    .btn-ghost{border:1px solid #1f2937;color:#6b7280;padding:8px 16px;border-radius:8px;
               font-size:13px;transition:all .15s;cursor:pointer}
    .btn-ghost:hover{border-color:#374151;color:#9ca3af}
    .btn-ghost:disabled{opacity:.4;cursor:not-allowed}
    @keyframes spin{to{transform:rotate(360deg)}}
    .spinner{width:20px;height:20px;border:2px solid #1f2937;border-top-color:#0B6E4F;
             border-radius:50%;animation:spin 1s linear infinite;display:inline-block}
  </style>
</head>
<body class="dm-grid-bg">
<div class="max-w-[1600px] mx-auto px-4 py-6">

  <!-- Header -->
  <header class="header-bar flex items-center gap-4">
    <!-- Logo icon — DataMatrix motif, exact match to landing page -->
    <a href="#" class="flex items-center gap-3">
      <div class="w-10 h-10 rounded-xl flex items-center justify-center shrink-0"
           style="background:#0B6E4F;box-shadow:0 0 24px rgba(11,110,79,0.45)">
        <svg width="26" height="26" viewBox="0 0 200 200" fill="none">
          <g fill="#FFFFFF">
            <rect x="22" y="22" width="18" height="18" rx="3"/>
            <rect x="160" y="22" width="18" height="18" rx="3"/>
            <rect x="22" y="160" width="18" height="18" rx="3"/>
            <rect x="60" y="56" width="20" height="88" rx="3"/>
            <rect x="60" y="56" width="64" height="20" rx="3"/>
            <rect x="104" y="56" width="20" height="44" rx="3"/>
            <rect x="60" y="88" width="60" height="16" rx="3"/>
            <rect x="92" y="108" width="16" height="16" rx="3"/>
            <rect x="108" y="124" width="16" height="20" rx="3"/>
          </g>
          <circle cx="168" cy="168" r="10" fill="#C4F542"/>
        </svg>
      </div>
      <div class="flex flex-col leading-none gap-0.5">
        <h1 class="font-display font-bold text-[16px] tracking-tight">
          ASL <span style="color:#C4F542">RAQAM</span>
        </h1>
        <span class="mono text-[9px] tracking-[0.18em] uppercase" style="color:#6B7A74">Проверка кодов маркировки</span>
      </div>
    </a>
    <div class="ml-auto mono text-xs border rounded px-2 py-0.5" style="color:#3A4A42;border-color:#1F2A26">v2.1</div>
  </header>

  <!-- Token Card -->
  <div class="card p-4 mb-5">
    <div class="flex items-center justify-between cursor-pointer select-none" onclick="toggleToken()">
      <div class="flex items-center gap-2">
        <div id="tok-dot" class="w-2 h-2 rounded-full bg-red-500"></div>
        <span class="text-sm font-medium text-gray-300">Bearer токен</span>
        <span id="tok-hint" class="text-xs text-gray-600 hidden">— сохранён</span>
      </div>
      <span id="tok-chevron" class="text-gray-600 text-xs">▼</span>
    </div>
    <div id="tok-body" class="mt-3 flex gap-2">
      <input id="tok-input" type="password" placeholder="Вставьте токен из личного кабинета xtrace.aslbelgisi.uz..."
             class="flex-1 bg-[#0e0e0e] border border-gray-800 rounded-lg px-3 py-2 text-sm mono
                    text-gray-200 placeholder-gray-700 transition-colors"/>
      <button class="btn-primary" onclick="saveToken()">Сохранить</button>
    </div>
  </div>

  <!-- Rate Limit Settings -->
  <div class="card p-3 mb-4">
    <div class="flex items-center justify-between cursor-pointer select-none" onclick="toggleSettings()">
      <div class="flex items-center gap-2">
        <span class="text-gray-600 text-sm">⚙</span>
        <span class="text-sm text-gray-400">Скорость запросов</span>
        <span id="rl-summary" class="text-xs mono text-gray-600 ml-1">— 300 мс / чанк</span>
      </div>
      <span id="rl-chevron" class="text-gray-700 text-xs">▶</span>
    </div>
    <div id="rl-body" class="hidden mt-3">
      <p class="text-xs text-gray-600 mb-2">
        Задержка между чанками (1 000 кодов = 1 чанк). При миллионах кодов увеличьте задержку,
        чтобы избежать блокировки со стороны ASL Belgisi.
      </p>
      <div class="flex flex-wrap gap-2 mb-3">
        <button class="filter-pill" data-ms="100"  onclick="setChunkDelay(100,this)">100 мс <span class="text-gray-700 ml-1">быстро</span></button>
        <button class="filter-pill active" data-ms="300"  onclick="setChunkDelay(300,this)">300 мс <span class="text-gray-700 ml-1">норма</span></button>
        <button class="filter-pill" data-ms="500"  onclick="setChunkDelay(500,this)">500 мс <span class="text-gray-700 ml-1">осторожно</span></button>
        <button class="filter-pill" data-ms="1000" onclick="setChunkDelay(1000,this)">1 сек <span class="text-gray-700 ml-1">безопасно</span></button>
        <button class="filter-pill" data-ms="2000" onclick="setChunkDelay(2000,this)">2 сек <span class="text-gray-700 ml-1">очень тихо</span></button>
      </div>
      <p class="text-xs text-gray-700">
        Сервер дополнительно выдерживает 300 мс между чанками и автоматически повторяет запрос при ошибке 429 (до 3 раз с паузой).
      </p>
    </div>
  </div>

  <!-- Tabs -->
  <div class="flex flex-wrap gap-1 bg-[#111] rounded-xl p-1 mb-5 border border-gray-900 w-fit">
    <button id="tab-btn-info" onclick="switchTab('info')"
            class="tab-btn tab-active px-4 py-2 rounded-lg text-sm transition-all text-gray-400">Инфо</button>
    <button id="tab-btn-history" onclick="switchTab('history')"
            class="tab-btn px-4 py-2 rounded-lg text-sm transition-all text-gray-400">История</button>
    <button id="tab-btn-docs" onclick="switchTab('docs')"
            class="tab-btn px-4 py-2 rounded-lg text-sm transition-all text-gray-400">Документы</button>
    <button id="tab-btn-export" onclick="switchTab('export')"
            class="tab-btn px-4 py-2 rounded-lg text-sm transition-all text-gray-400">Выгрузка КМ</button>
    <button id="tab-btn-reconcile" onclick="switchTab('reconcile')"
            class="tab-btn px-4 py-2 rounded-lg text-sm transition-all text-gray-400">Сверка</button>
  </div>

  <!-- ═══════════ INFO PANEL ═══════════ -->
  <div id="panel-info">

    <!-- Input + File Upload -->
    <div class="card p-4 mb-4">
      <div class="flex items-center justify-between mb-2">
        <label class="text-xs text-gray-600 uppercase tracking-widest">Коды маркировки</label>
        <div class="flex items-center gap-3">
          <span id="code-count" class="text-xs mono text-gray-600">0 кодов</span>
          <!-- File upload -->
          <label class="btn-ghost text-xs px-3 py-1.5 cursor-pointer" style="padding:5px 10px">
            ↑ CSV / TXT
            <input id="file-input" type="file" accept=".csv,.txt" class="hidden" onchange="loadFile(this)"/>
          </label>
        </div>
      </div>
      <textarea id="code-input" rows="6" oninput="onCodeInput()"
                placeholder="Вставьте коды — каждый код на отдельной строке (или загрузите CSV/TXT файл выше)..."
                class="w-full bg-[#0a0a0a] border border-gray-800 rounded-lg p-3 text-sm mono
                       text-gray-200 placeholder-gray-700 resize-y transition-colors"></textarea>
      <p class="text-xs text-gray-700 mt-1.5">
        Каждый код на отдельной строке &nbsp;·&nbsp; Макс. 50 000 кодов
      </p>
    </div>

    <!-- Actions -->
    <div class="flex flex-wrap gap-2 mb-5">
      <button class="btn-primary" onclick="runInfo()" id="btn-check">Проверить</button>
      <button class="btn-ghost" onclick="resetInfo()">Сбросить</button>
      <div class="flex-1"></div>
      <button class="btn-ghost" onclick="exportCSV()" id="btn-csv" disabled>↓ CSV</button>
      <button class="btn-ghost" onclick="exportJSON()" id="btn-json" disabled>↓ JSON</button>
    </div>

    <!-- Progress -->
    <div id="progress-wrap" class="hidden mb-4">
      <div class="flex justify-between text-xs text-gray-600 mb-1.5">
        <span>Загрузка данных...</span>
        <span id="progress-text" class="mono">0 / 0</span>
      </div>
      <div class="w-full h-1 bg-gray-900 rounded-full overflow-hidden">
        <div id="progress-bar" class="h-full rounded-full bar-pulse transition-all"
             style="background:#0B6E4F;width:0%"></div>
      </div>
    </div>

    <!-- Error -->
    <div id="err-banner" class="hidden border border-red-900 bg-red-950/30 rounded-xl p-4 mb-4">
      <p class="text-red-400 text-sm" id="err-text"></p>
    </div>

    <!-- Summary -->
    <div id="summary" class="hidden grid grid-cols-4 gap-3 mb-5">
      <div class="card p-3 text-center">
        <div class="text-3xl font-display font-bold text-white" id="s-total">0</div>
        <div class="text-xs text-gray-600 mt-0.5">Всего</div>
      </div>
      <div class="card p-3 text-center">
        <div class="text-3xl font-display font-bold" style="color:#22c55e" id="s-ok">0</div>
        <div class="text-xs text-gray-600 mt-0.5">В обороте</div>
      </div>
      <div class="card p-3 text-center">
        <div class="text-3xl font-display font-bold" style="color:#f59e0b" id="s-other">0</div>
        <div class="text-xs text-gray-600 mt-0.5">Другой статус</div>
      </div>
      <div class="card p-3 text-center">
        <div class="text-3xl font-display font-bold" style="color:#ef4444" id="s-err">0</div>
        <div class="text-xs text-gray-600 mt-0.5">Не найдено</div>
      </div>
    </div>

    <!-- Status Filters -->
    <div id="filter-wrap" class="hidden flex flex-wrap gap-2 mb-3">
      <button class="filter-pill active" data-filter="all"        onclick="setFilter('all')">Все</button>
      <button class="filter-pill" data-filter="INTRODUCED"        onclick="setFilter('INTRODUCED')">В обороте</button>
      <button class="filter-pill" data-filter="APPLIED"           onclick="setFilter('APPLIED')">Нанесён</button>
      <button class="filter-pill" data-filter="RECEIVED"          onclick="setFilter('RECEIVED')">Получен</button>
      <button class="filter-pill" data-filter="WITHDRAWN"         onclick="setFilter('WITHDRAWN')">Выведен</button>
      <button class="filter-pill" data-filter="WRITTEN_OFF"       onclick="setFilter('WRITTEN_OFF')">Списан</button>
      <button class="filter-pill" data-filter="NOT_FOUND"         onclick="setFilter('NOT_FOUND')">Не найден</button>
    </div>

    <!-- Table -->
    <div id="result-wrap" class="hidden">
      <div class="card overflow-hidden">
        <div class="overflow-x-auto">
          <table class="w-full text-sm border-collapse">
            <thead>
              <tr style="background:#0f0f0f;border-bottom:1px solid #1a1a1a">
                <th class="w-8 px-3 py-3">
                  <input type="checkbox" id="chk-all" onchange="toggleAll()">
                </th>
                <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">Код (CIS)</th>
                <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">GTIN</th>
                <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">Тип упаковки</th>
                <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">Годен до</th>
                <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">ИНН эмитента</th>
                <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">Статус</th>
                <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">АИК</th>
              </tr>
            </thead>
            <tbody id="result-body" class="divide-y divide-gray-900/60"></tbody>
          </table>
        </div>
      </div>
      <p class="text-xs text-gray-700 mt-2 text-right mono" id="result-label"></p>
    </div>

  </div>

  <!-- ═══════════ HISTORY PANEL ═══════════ -->
  <div id="panel-history" class="hidden">
    <div class="card p-4 mb-4">
      <label class="text-xs text-gray-600 uppercase tracking-widest block mb-2">Код маркировки</label>
      <div class="flex gap-2">
        <input id="history-input" type="text" placeholder="Введите один код..."
               class="flex-1 bg-[#0a0a0a] border border-gray-800 rounded-lg px-3 py-2 text-sm mono
                      text-gray-200 placeholder-gray-700 transition-colors"
               onkeydown="if(event.key==='Enter') runHistory()"/>
        <button class="btn-primary" onclick="runHistory()">Получить историю</button>
      </div>
    </div>
    <div id="history-loading" class="hidden text-center py-8">
      <div class="spinner"></div>
      <p class="text-xs text-gray-600 mt-2">Загрузка...</p>
    </div>
    <div id="history-results"></div>
  </div>

  <!-- ═══════════ DOCS PANEL ═══════════ -->
  <div id="panel-docs" class="hidden">

    <!-- Sub-tabs -->
    <div class="flex gap-1 bg-[#111] rounded-xl p-1 mb-4 border border-gray-900 w-fit">
      <button id="subtab-btn-doclist" onclick="switchDocsSubtab('doclist')"
              class="subtab-btn subtab-active px-4 py-1.5 rounded-lg text-xs transition-all text-gray-400">Документы (ЭСФ, Заказы...)</button>
      <button id="subtab-btn-receipts" onclick="switchDocsSubtab('receipts')"
              class="subtab-btn px-4 py-1.5 rounded-lg text-xs transition-all text-gray-400">Чеки ККТ</button>
    </div>

    <!-- ── SUB-PANEL: Документы ── -->
    <div id="subpanel-doclist">

    <!-- Filters -->
    <div class="card p-4 mb-4">
      <div class="grid grid-cols-1 md:grid-cols-2 gap-4">
        <div>
          <p class="text-xs text-gray-600 uppercase tracking-widest mb-2">Статус документа</p>
          <div class="flex flex-wrap gap-2" id="doc-status-pills">
            <button class="filter-pill active" data-val="SUCCESS"             onclick="toggleDocFilter('status','SUCCESS',this)">Успешно</button>
            <button class="filter-pill active" data-val="ERROR"               onclick="toggleDocFilter('status','ERROR',this)">Ошибка</button>
            <button class="filter-pill active" data-val="PARTIALLY_PROCESSED" onclick="toggleDocFilter('status','PARTIALLY_PROCESSED',this)">Частично обработан</button>
            <button class="filter-pill active" data-val="IN_PROCESS"          onclick="toggleDocFilter('status','IN_PROCESS',this)">В обработке</button>
          </div>
        </div>
        <div>
          <p class="text-xs text-gray-600 uppercase tracking-widest mb-2">Тип документа (ЭСФ)</p>
          <div id="doc-type-pills" class="flex flex-wrap gap-2">
            <button class="filter-pill active" data-val="ORIGINAL_INVOICE"   onclick="toggleDocFilter('types','ORIGINAL_INVOICE',this)">ЭСФ стандартная</button>
            <button class="filter-pill active" data-val="ADDITIONAL_INVOICE" onclick="toggleDocFilter('types','ADDITIONAL_INVOICE',this)">ЭСФ дополнительная</button>
            <button class="filter-pill active" data-val="CORRECTED_INVOICE"  onclick="toggleDocFilter('types','CORRECTED_INVOICE',this)">ЭСФ исправленная</button>
            <button class="filter-pill" data-val="CANCELLED_INVOICE"         onclick="toggleDocFilter('types','CANCELLED_INVOICE',this)">ЭСФ аннулированная</button>
          </div>
        </div>
      </div>
      <div class="grid grid-cols-2 gap-3 mt-3">
        <div>
          <label class="text-xs text-gray-600 block mb-1">Дата от</label>
          <input type="date" id="doc-date-from" class="w-full bg-[#0a0a0a] border border-gray-800 rounded-lg px-3 py-2 text-sm text-gray-300 transition-colors"/>
        </div>
        <div>
          <label class="text-xs text-gray-600 block mb-1">Дата до</label>
          <input type="date" id="doc-date-to" class="w-full bg-[#0a0a0a] border border-gray-800 rounded-lg px-3 py-2 text-sm text-gray-300 transition-colors"/>
        </div>
      </div>
      <div class="flex gap-2 mt-3 flex-wrap items-center">
        <button class="btn-primary" onclick="loadDocs()">Найти документы</button>
        <button class="btn-ghost" onclick="resetDocs()">Сбросить</button>
        <button class="btn-ghost" id="btn-docs-excel" onclick="exportDocsExcel()" disabled>↓ Список Excel</button>
        <button class="btn-ghost" id="btn-bulk-codes" onclick="downloadAllDocCodes()" disabled
                title="Скачать все коды маркировки из всех загруженных ЭСФ одним Excel-файлом">↓ Все КМ из ЭСФ</button>
        <span id="docs-count-label" class="text-xs text-gray-600 ml-2"></span>
      </div>
      <div id="bulk-progress-wrap" class="hidden mt-2 flex items-center gap-2">
        <div class="spinner" style="width:14px;height:14px;border-width:2px"></div>
        <span id="bulk-progress-text" class="text-xs text-gray-500"></span>
      </div>
      <!-- Прямой поиск по ID -->
      <div class="mt-3 pt-3 border-t border-gray-900 flex gap-2">
        <input id="doc-direct-id" type="text" placeholder="Поиск по ID документа (чека ККТ или любого другого)..."
               class="flex-1 bg-[#0a0a0a] border border-gray-800 rounded-lg px-3 py-2 text-xs mono
                      text-gray-200 placeholder-gray-700 transition-colors"
               onkeydown="if(event.key==='Enter') loadDocById()"/>
        <button class="btn-ghost text-xs" onclick="loadDocById()">Открыть</button>
      </div>
    </div>

    <div id="docs-loading" class="hidden text-center py-6"><div class="spinner"></div></div>
    <div id="docs-err" class="hidden border border-red-900 bg-red-950/30 rounded-xl p-4 mb-4">
      <p class="text-red-400 text-sm" id="docs-err-text"></p>
    </div>
    <div id="docs-diag-wrap" class="hidden border border-yellow-900/40 bg-yellow-950/20 rounded-xl p-3 mb-3">
      <p class="text-yellow-600 text-xs mono" id="docs-diag"></p>
    </div>

    <!-- Docs Table -->
    <div id="docs-result-wrap" class="hidden">
      <div class="card overflow-hidden">
        <div class="overflow-x-auto">
          <table class="w-full text-sm border-collapse">
            <thead>
              <tr style="background:#0f0f0f;border-bottom:1px solid #1a1a1a">
                <th class="w-8 px-3 py-3"><input type="checkbox" id="docs-chk-all" onchange="toggleDocsAll()"></th>
                <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">Номер ЭСФ</th>
                <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">FacturaId</th>
                <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">Тип</th>
                <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">Статус</th>
                <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider cursor-pointer select-none hover:text-gray-400 transition-colors"
                    onclick="sortDocsByDate()" id="docs-date-th">Дата ↕</th>
              </tr>
            </thead>
            <tbody id="docs-tbody" class="divide-y divide-gray-900/60"></tbody>
          </table>
        </div>
      </div>
      <div id="docs-load-more" class="hidden mt-3 text-center py-3">
        <div class="spinner" style="display:inline-block"></div>
        <p class="text-xs text-gray-600 mt-1">Загружаю следующую страницу...</p>
      </div>
    </div>

    </div><!-- /subpanel-doclist -->

    <!-- ── SUB-PANEL: Чеки ККТ ── -->
    <div id="subpanel-receipts" class="hidden">

      <!-- Sync instruction card -->
      <div class="card p-4 mb-4" style="border-color:#1a2a1a">
        <div class="flex items-center justify-between mb-3">
          <p class="text-sm text-gray-300 font-semibold">Синхронизация чеков ККТ</p>
          <div class="text-right">
            <div id="rcp-sync-status" class="text-xs text-gray-600">Нет данных</div>
            <div id="rcp-sync-time" class="text-xs mono text-gray-700 mt-0.5"></div>
          </div>
        </div>

        <!-- Steps -->
        <div class="space-y-3">
          <!-- Step 1 -->
          <div class="flex gap-3 items-start">
            <div class="w-6 h-6 rounded-full flex items-center justify-center shrink-0 text-xs font-bold" style="background:#0B6E4F;color:#C4F542">1</div>
            <div class="flex-1">
              <p class="text-xs text-gray-300">Войдите в личный кабинет через ЭЦП</p>
              <a href="https://xtrace.aslbelgisi.uz" target="_blank"
                 class="text-xs mono mt-1 inline-block" style="color:#0B6E4F">xtrace.aslbelgisi.uz ↗</a>
            </div>
          </div>

          <!-- Step 2 -->
          <div class="flex gap-3 items-start">
            <div class="w-6 h-6 rounded-full flex items-center justify-center shrink-0 text-xs font-bold" style="background:#0B6E4F;color:#C4F542">2</div>
            <div class="flex-1">
              <p class="text-xs text-gray-300 mb-1">
                Нажмите <kbd class="px-1.5 py-0.5 rounded text-gray-400" style="background:#1a1a1a;border:1px solid #333;font-size:11px">F12</kbd>
                → вкладка <span class="text-gray-400">«Network»</span>
                → нажмите «Чеки ККТ» в ЛК (или обновите страницу)
                → выберите любой запрос к <span class="mono text-gray-400">/api/rcp/</span>
                → вкладка <span class="text-gray-400">«Headers»</span>
                → скопируйте значение <span class="mono text-gray-400">Authorization</span>
                <span class="text-gray-600">(без слова Bearer, только UUID)</span>
              </p>
              <div class="flex gap-2 items-center mt-1 p-2 rounded" style="background:#0a0a0a;border:1px solid #1f2937">
                <span class="text-gray-600 text-xs shrink-0">Bearer</span>
                <input id="rcp-session-token" type="text"
                       placeholder="f5ff59af-7fc7-4b51-aa67-6defcbba961e"
                       class="flex-1 bg-transparent border-none mono text-xs text-green-400 placeholder-gray-700 focus:outline-none"/>
                <button onclick="syncRcpWithToken()" class="btn-primary text-xs shrink-0" style="padding:4px 12px">Синхронизировать</button>
              </div>
              <div id="rcp-sync-progress" class="hidden mt-2 flex items-center gap-2">
                <div class="spinner" style="width:14px;height:14px;border-width:2px"></div>
                <span class="text-xs text-gray-500">Загружаю все чеки...</span>
              </div>
              <div id="rcp-sync-ok" class="hidden mt-2 text-xs" style="color:#22c55e"></div>
            </div>
          </div>

          <!-- Step 3 -->
          <div class="flex gap-3 items-start">
            <div class="w-6 h-6 rounded-full flex items-center justify-center shrink-0 text-xs font-bold" style="background:#0B6E4F;color:#C4F542">3</div>
            <div class="flex-1">
              <p class="text-xs text-gray-300">После синхронизации нажмите кнопку <span class="text-gray-400">«Загрузить»</span> ниже</p>
            </div>
          </div>
        </div>
      </div>

      <!-- Filters + actions -->
      <div class="card p-4 mb-4">
        <div class="grid grid-cols-2 gap-3 mb-3">
          <div>
            <label class="text-xs text-gray-600 block mb-1">Дата от</label>
            <input type="date" id="rcp-date-from" class="w-full bg-[#0a0a0a] border border-gray-800 rounded-lg px-3 py-2 text-sm text-gray-300 transition-colors"/>
          </div>
          <div>
            <label class="text-xs text-gray-600 block mb-1">Дата до</label>
            <input type="date" id="rcp-date-to" class="w-full bg-[#0a0a0a] border border-gray-800 rounded-lg px-3 py-2 text-sm text-gray-300 transition-colors"/>
          </div>
        </div>
        <div class="mb-3">
          <p class="text-xs text-gray-600 uppercase tracking-widest mb-2">Тип чека</p>
          <div class="flex gap-2">
            <button class="filter-pill active" data-rcp-type="SALE"   onclick="toggleRcpType('SALE',this)">Продажа</button>
            <button class="filter-pill active" data-rcp-type="RETURN" onclick="toggleRcpType('RETURN',this)">Возврат</button>
          </div>
        </div>
        <div class="mb-3">
          <label class="text-xs text-gray-600 block mb-1">Поиск по внешнему ID</label>
          <input type="text" id="rcp-ext-id-filter" placeholder="LG420211631..." oninput="applyRcpFilter()"
                 class="w-full bg-[#0a0a0a] border border-gray-800 rounded-lg px-3 py-2 text-xs mono text-gray-200 placeholder-gray-700 transition-colors"/>
        </div>
        <div class="flex gap-2 flex-wrap items-center">
          <button class="btn-primary" onclick="applyRcpFilter()">Загрузить</button>
          <button class="btn-ghost" onclick="resetRcpFilter()">Сбросить фильтры</button>
          <button class="btn-ghost" id="btn-rcp-excel" onclick="exportReceiptsExcel()" disabled>↓ Excel</button>
          <button class="btn-ghost" id="btn-rcp-bulk" onclick="downloadAllRcpCodes()" disabled>↓ Все КМ из чеков</button>
          <span id="rcp-count-label" class="text-xs text-gray-600 ml-2"></span>
        </div>
        <div id="rcp-bulk-progress-wrap" class="hidden mt-3 flex items-center gap-3">
          <div class="spinner" style="width:14px;height:14px;border-width:2px"></div>
          <span id="rcp-bulk-progress-text" class="text-xs text-gray-500">Собираю коды...</span>
        </div>
      </div>

      <div id="rcp-loading" class="hidden text-center py-6"><div class="spinner"></div></div>
      <div id="rcp-err" class="hidden border border-red-900 bg-red-950/30 rounded-xl p-4 mb-4">
        <p class="text-red-400 text-sm" id="rcp-err-text"></p>
      </div>

      <div id="rcp-result-wrap" class="hidden">
        <div class="card overflow-hidden">
          <div class="overflow-x-auto">
            <table class="w-full text-sm border-collapse">
              <thead>
                <tr style="background:#0f0f0f;border-bottom:1px solid #1a1a1a">
                  <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">ID чека</th>
                  <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">Внешний ID</th>
                  <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">Тип</th>
                  <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">Статус</th>
                  <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">Дата регистрации</th>
                  <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">Продавец</th>
                </tr>
              </thead>
              <tbody id="rcp-tbody" class="divide-y divide-gray-900/60"></tbody>
            </table>
          </div>
        </div>
      </div>
    </div><!-- /subpanel-receipts -->

  </div>

  <!-- ═══════════ EXPORT PANEL ═══════════ -->
  <div id="panel-export" class="hidden">
    <div class="card p-4 mb-4">
      <div class="grid grid-cols-1 md:grid-cols-2 gap-4">
        <div>
          <label class="text-xs text-gray-600 uppercase tracking-widest block mb-1">
            GTIN <span class="text-red-500 normal-case">* обязательно</span>
          </label>
          <input type="text" id="exp-gtin" placeholder="04780161040249..."
                 class="w-full bg-[#0a0a0a] border border-gray-800 rounded-lg px-3 py-2 text-sm mono text-gray-200 placeholder-gray-700 transition-colors"/>
          <p class="text-xs text-gray-700 mt-1">API требует точный GTIN для выгрузки</p>
        </div>
        <div>
          <p class="text-xs text-gray-600 uppercase tracking-widest mb-2">Тип упаковки</p>
          <div class="flex flex-wrap gap-2">
            <button class="filter-pill active" data-val="UNIT"     onclick="toggleExpFilter('pkg','UNIT',this)">Единица</button>
            <button class="filter-pill" data-val="GROUP"           onclick="toggleExpFilter('pkg','GROUP',this)">Группа</button>
            <button class="filter-pill" data-val="BOX_LV_1"        onclick="toggleExpFilter('pkg','BOX_LV_1',this)">Короб</button>
            <button class="filter-pill" data-val="BOX_LV_2"        onclick="toggleExpFilter('pkg','BOX_LV_2',this)">Паллет</button>
          </div>
        </div>
      </div>
      <div class="mt-3">
        <p class="text-xs text-gray-600 uppercase tracking-widest mb-2">Статус кодов</p>
        <div class="flex flex-wrap gap-2">
          <button class="filter-pill active" data-val="INTRODUCED"  onclick="toggleExpFilter('status','INTRODUCED',this)">В обороте</button>
          <button class="filter-pill" data-val="WITHDRAWN"          onclick="toggleExpFilter('status','WITHDRAWN',this)">Выведен</button>
          <button class="filter-pill" data-val="APPLIED"            onclick="toggleExpFilter('status','APPLIED',this)">Нанесён</button>
          <button class="filter-pill" data-val="RECEIVED"           onclick="toggleExpFilter('status','RECEIVED',this)">Получен</button>
          <button class="filter-pill" data-val="WRITTEN_OFF"        onclick="toggleExpFilter('status','WRITTEN_OFF',this)">Списан</button>
        </div>
      </div>
      <div class="grid grid-cols-2 md:grid-cols-4 gap-3 mt-3">
        <div><label class="text-xs text-gray-600 block mb-1">Эмиссия от</label>
          <input type="date" id="exp-em-from" class="w-full bg-[#0a0a0a] border border-gray-800 rounded-lg px-3 py-2 text-sm text-gray-300 transition-colors"/></div>
        <div><label class="text-xs text-gray-600 block mb-1">Эмиссия до</label>
          <input type="date" id="exp-em-to" class="w-full bg-[#0a0a0a] border border-gray-800 rounded-lg px-3 py-2 text-sm text-gray-300 transition-colors"/></div>
        <div><label class="text-xs text-gray-600 block mb-1">Годен от</label>
          <input type="date" id="exp-ex-from" class="w-full bg-[#0a0a0a] border border-gray-800 rounded-lg px-3 py-2 text-sm text-gray-300 transition-colors"/></div>
        <div><label class="text-xs text-gray-600 block mb-1">Годен до</label>
          <input type="date" id="exp-ex-to" class="w-full bg-[#0a0a0a] border border-gray-800 rounded-lg px-3 py-2 text-sm text-gray-300 transition-colors"/></div>
      </div>
      <div class="mt-4">
        <button class="btn-primary" onclick="startExport()">Создать задание на выгрузку</button>
      </div>
    </div>

    <!-- Job Status -->
    <div id="exp-job-wrap" class="hidden card p-4">
      <div class="flex items-center justify-between mb-3">
        <div>
          <p class="text-xs text-gray-600">Задание</p>
          <p class="mono text-xs text-gray-400 mt-0.5" id="exp-job-id">—</p>
        </div>
        <div class="flex items-center gap-3">
          <span class="text-xs text-gray-700">⏱ <span id="exp-elapsed">0 сек</span></span>
          <span id="exp-job-badge" class="status-badge" style="background:#111;color:#6b7280;border-color:#1f2937">—</span>
        </div>
      </div>
      <!-- Steps -->
      <div class="flex items-center gap-2 text-xs mb-3">
        <span id="step-pending"    class="px-2 py-1 rounded border border-gray-800 text-gray-600">Ожидание</span>
        <span class="text-gray-800">→</span>
        <span id="step-processing" class="px-2 py-1 rounded border border-gray-800 text-gray-600">Обработка</span>
        <span class="text-gray-800">→</span>
        <span id="step-done"       class="px-2 py-1 rounded border border-gray-800 text-gray-600">Готово</span>
      </div>
      <div class="flex items-center gap-3 mb-2">
        <button id="exp-download-btn" class="hidden btn-primary" onclick="downloadExport()">↓ Скачать ZIP</button>
        <button id="exp-refresh-btn" class="btn-ghost text-xs" onclick="checkExportNow()">↻ Проверить сейчас</button>
        <span class="text-xs text-gray-700">Опрос #<span id="exp-poll-count">0</span> · сырой статус: <span id="exp-raw-status" class="mono">—</span></span>
      </div>
      <p class="text-xs text-gray-700" id="exp-job-hint"></p>
    </div>

    <div id="exp-err" class="hidden border border-red-900 bg-red-950/30 rounded-xl p-4 mt-4">
      <p class="text-red-400 text-sm" id="exp-err-text"></p>
    </div>
  </div>

  <!-- ═══════════ RECONCILE PANEL ═══════════ -->
  <div id="panel-reconcile" class="hidden">

    <div class="grid grid-cols-1 md:grid-cols-3 gap-4 mb-4">
      <div class="md:col-span-2 card p-4">
        <div class="flex items-center justify-between mb-2">
          <label class="text-xs text-gray-600 uppercase tracking-widest">Коды для сверки (из ЭСФ / накладной)</label>
          <div class="flex items-center gap-3">
            <span id="rec-count" class="text-xs mono text-gray-600">0 кодов</span>
            <label class="btn-ghost text-xs cursor-pointer" style="padding:5px 10px">
              ↑ Файл
              <input type="file" accept=".csv,.txt" class="hidden" onchange="loadRecFile(this)"/>
            </label>
          </div>
        </div>
        <textarea id="rec-input" rows="6" oninput="onRecInput()"
                  placeholder="Каждый код на отдельной строке..."
                  class="w-full bg-[#0a0a0a] border border-gray-800 rounded-lg p-3 text-sm mono
                         text-gray-200 placeholder-gray-700 resize-y transition-colors"></textarea>
      </div>
      <div class="card p-4">
        <label class="text-xs text-gray-600 uppercase tracking-widest block mb-2">ИНН вашей организации</label>
        <input type="text" id="rec-inn" placeholder="123456789..."
               class="w-full bg-[#0a0a0a] border border-gray-800 rounded-lg px-3 py-2 text-sm mono
                      text-gray-200 placeholder-gray-700 transition-colors mb-3"/>
        <p class="text-xs text-gray-700 leading-relaxed">
          Коды, у которых <span class="text-gray-500">issuerTin ≠ ваш ИНН</span>, будут отмечены как «Не ваш код».
        </p>
      </div>
    </div>

    <div class="flex flex-wrap gap-2 mb-5">
      <button class="btn-primary" onclick="runReconcile()" id="btn-rec">Проверить сверку</button>
      <button class="btn-ghost" onclick="resetReconcile()">Сбросить</button>
      <div class="flex-1"></div>
      <button class="btn-ghost" onclick="exportRecCSV()" id="btn-rec-csv" disabled>↓ Проблемные CSV</button>
      <button class="btn-ghost" onclick="exportRecJSON()" id="btn-rec-json" disabled>↓ Все JSON</button>
    </div>

    <div id="rec-progress-wrap" class="hidden mb-4">
      <div class="flex justify-between text-xs text-gray-600 mb-1.5">
        <span>Проверка кодов...</span><span id="rec-progress-text" class="mono">0 / 0</span>
      </div>
      <div class="w-full h-1 bg-gray-900 rounded-full overflow-hidden">
        <div id="rec-progress-bar" class="h-full rounded-full bar-pulse" style="background:#0B6E4F;width:0%"></div>
      </div>
    </div>

    <div id="rec-err" class="hidden border border-red-900 bg-red-950/30 rounded-xl p-4 mb-4">
      <p class="text-red-400 text-sm" id="rec-err-text"></p>
    </div>

    <!-- Summary -->
    <div id="rec-summary" class="hidden grid grid-cols-4 gap-3 mb-5">
      <div class="card p-3 text-center"><div class="text-3xl font-display font-bold text-white" id="rs-total">0</div><div class="text-xs text-gray-600 mt-0.5">Всего</div></div>
      <div class="card p-3 text-center"><div class="text-3xl font-display font-bold" style="color:#22c55e" id="rs-ok">0</div><div class="text-xs text-gray-600 mt-0.5">В обороте</div></div>
      <div class="card p-3 text-center"><div class="text-3xl font-display font-bold" style="color:#f59e0b" id="rs-warn">0</div><div class="text-xs text-gray-600 mt-0.5">Проблемы</div></div>
      <div class="card p-3 text-center"><div class="text-3xl font-display font-bold" style="color:#6b7280" id="rs-miss">0</div><div class="text-xs text-gray-600 mt-0.5">Не найден</div></div>
    </div>

    <!-- Filter pills -->
    <div id="rec-filter-wrap" class="hidden flex flex-wrap gap-2 mb-3">
      <button class="filter-pill active" data-filter="all"      onclick="setRecFilter('all')">Все</button>
      <button class="filter-pill" data-filter="ok"              onclick="setRecFilter('ok')">✅ В обороте</button>
      <button class="filter-pill" data-filter="status"          onclick="setRecFilter('status')">⚠️ Статус не тот</button>
      <button class="filter-pill" data-filter="owner"           onclick="setRecFilter('owner')">❌ Не ваш код</button>
      <button class="filter-pill" data-filter="missing"         onclick="setRecFilter('missing')">❓ Не найден</button>
    </div>

    <!-- Table -->
    <div id="rec-result-wrap" class="hidden">
      <div class="card overflow-hidden">
        <div class="overflow-x-auto">
          <table class="w-full text-sm border-collapse">
            <thead>
              <tr style="background:#0f0f0f;border-bottom:1px solid #1a1a1a">
                <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">Код</th>
                <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">Статус API</th>
                <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">ИНН эмитента</th>
                <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">Итог</th>
                <th class="text-left px-3 py-3 text-xs text-gray-600 uppercase tracking-wider">Годен до</th>
              </tr>
            </thead>
            <tbody id="rec-tbody" class="divide-y divide-gray-900/60"></tbody>
          </table>
        </div>
      </div>
      <p class="text-xs text-gray-700 mt-2 text-right mono" id="rec-result-label"></p>
    </div>

  </div>

</div>

<script>
// ── State ──────────────────────────────────────────────────────────────────
const S = {
  token: '',
  chunkDelay: 300,   // ms between chunks (rate-limit guard)
  // Info tab
  allResults: [],
  activeFilter: 'all',
  selectedCis: new Set(),
  // Docs tab
  docs: [],
  docsCursor: null,
  docsStatusFilter: new Set(['SUCCESS','ERROR','PARTIALLY_PROCESSED','IN_PROCESS']),
  docsTypesFilter: new Set(['ORIGINAL_INVOICE','ADDITIONAL_INVOICE','CORRECTED_INVOICE']),
  // Export tab
  expStatusFilter: new Set(['INTRODUCED']),
  expPkgFilter: new Set(['UNIT']),
  exportJobId: null,
  exportPollTimer: null,
  exportElapsedTimer: null,
  exportPollCount: 0,
  // Reconcile tab
  reconcileResults: [],
  reconcileFilter: 'all',
};

// Status config: maps API status values → display
const STATUS = {
  INTRODUCED:  { label:'В обороте',  bg:'#052e1a', color:'#22c55e', border:'#14532d' },
  APPLIED:     { label:'Нанесён',    bg:'#0c1a2e', color:'#60a5fa', border:'#1e3a5f' },
  RECEIVED:    { label:'Получен',    bg:'#1c1400', color:'#fbbf24', border:'#451a03' },
  WITHDRAWN:   { label:'Выведен',    bg:'#1c0a0a', color:'#f87171', border:'#450a0a' },
  WRITTEN_OFF: { label:'Списан',     bg:'#111',    color:'#6b7280', border:'#1f2937' },
};

const EVENT_TYPES = {
  OWNER_CHANGE:     'Смена владельца',
  INTRODUCTION:     'Ввод в оборот',
  UTILISATION:      'Нанесение',
  VALIDATION:       'Валидация',
  WITHDRAWAL:       'Вывод из оборота',
  REVERSAL:         'Отмена',
  CHANGE_PARENT:    'Изменение родителя',
  CUSTOMS_CONTROL:  'Таможенный контроль',
};

// ── Init ───────────────────────────────────────────────────────────────────
document.addEventListener('DOMContentLoaded', () => {
  S.token = localStorage.getItem('asl_tok') || '';
  if (S.token) {
    document.getElementById('tok-input').value = S.token;
    setTokenUI(true);
    collapseToken();
  }

  // Restore chunk delay setting
  const savedDelay = parseInt(localStorage.getItem('asl_chunk_delay') || '300');
  S.chunkDelay = savedDelay;
  const delayBtn = document.querySelector(`#rl-body .filter-pill[data-ms="${savedDelay}"]`);
  if (delayBtn) {
    document.querySelectorAll('#rl-body .filter-pill').forEach(el => el.classList.remove('active'));
    delayBtn.classList.add('active');
    document.getElementById('rl-summary').textContent = `— ${savedDelay < 1000 ? savedDelay + ' мс' : (savedDelay/1000) + ' сек'} / чанк`;
  }

  switchTab('info');
});

// ── Token ──────────────────────────────────────────────────────────────────
function saveToken() {
  const v = document.getElementById('tok-input').value.trim();
  S.token = v;
  if (v) localStorage.setItem('asl_tok', v);
  setTokenUI(!!v);
  if (v) collapseToken();
}
function setTokenUI(ok) {
  document.getElementById('tok-dot').className = `w-2 h-2 rounded-full ${ok?'bg-emerald-500':'bg-red-500'}`;
  document.getElementById('tok-hint').classList.toggle('hidden', !ok);
}
function collapseToken() {
  document.getElementById('tok-body').classList.add('hidden');
  document.getElementById('tok-chevron').textContent = '▶';
}
function toggleToken() {
  const b = document.getElementById('tok-body');
  b.classList.toggle('hidden');
  document.getElementById('tok-chevron').textContent = b.classList.contains('hidden') ? '▶' : '▼';
}

function toggleSettings() {
  const b = document.getElementById('rl-body');
  b.classList.toggle('hidden');
  document.getElementById('rl-chevron').textContent = b.classList.contains('hidden') ? '▶' : '▼';
}

function setChunkDelay(ms, btn) {
  S.chunkDelay = ms;
  localStorage.setItem('asl_chunk_delay', ms);
  document.querySelectorAll('#rl-body .filter-pill').forEach(el => el.classList.remove('active'));
  btn.classList.add('active');
  document.getElementById('rl-summary').textContent = `— ${ms < 1000 ? ms + ' мс' : (ms/1000) + ' сек'} / чанк`;
}

// ── Tabs ───────────────────────────────────────────────────────────────────
const ALL_TABS = ['info','history','docs','export','reconcile'];
function switchTab(name) {
  ALL_TABS.forEach(t => {
    document.getElementById(`tab-btn-${t}`).classList.toggle('tab-active', t===name);
    document.getElementById(`panel-${t}`).classList.toggle('hidden', t!==name);
  });
}

// ── Code Input ─────────────────────────────────────────────────────────────
function parseCodes(raw) {
  // One code per line — do not split by spaces or commas,
  // codes may contain spaces, commas and other special chars
  return raw.split(/\r?\n/).map(s => s.trim()).filter(s => s.length > 0);
}

function onCodeInput() {
  const codes = parseCodes(document.getElementById('code-input').value);
  document.getElementById('code-count').textContent = `${codes.length} кодов`;
}

// ── CSV / TXT File Upload ──────────────────────────────────────────────────
function loadFile(input) {
  const file = input.files[0];
  if (!file) return;
  const reader = new FileReader();
  reader.onload = e => {
    const codes = e.target.result
      .split(/\r?\n/)
      .map(s => s.trim())
      .filter(s => s.length > 0);

    const unique = [...new Set(codes)];
    document.getElementById('code-input').value = unique.join('\n');
    document.getElementById('code-count').textContent = `${unique.length} кодов`;
    input.value = '';
  };
  reader.readAsText(file, 'UTF-8');
}

// ── Run Info ───────────────────────────────────────────────────────────────
async function runInfo() {
  if (!S.token) { showErr('Введите и сохраните токен.'); return; }
  const raw = document.getElementById('code-input').value.trim();
  if (!raw) return;

  const codes = [...new Set(parseCodes(raw))];
  if (!codes.length) { showErr('Не удалось распознать коды. Проверьте формат ввода.'); return; }
  if (codes.length > 50000) { showErr('Максимум 50 000 кодов за раз.'); return; }

  const chunks = [];
  for (let i = 0; i < codes.length; i += 1000) chunks.push(codes.slice(i, i+1000));

  S.allResults = [];
  S.selectedCis = new Set();
  S.activeFilter = 'all';
  hideErr();
  clearTable();
  setActiveFilterPill('all');
  document.getElementById('progress-wrap').classList.remove('hidden');
  document.getElementById('summary').classList.remove('hidden');
  document.getElementById('filter-wrap').classList.remove('hidden');
  document.getElementById('result-wrap').classList.remove('hidden');
  document.getElementById('btn-check').disabled = true;
  resetSummary();
  updateProgress(0, chunks.length);

  let done = 0;
  for (const chunk of chunks) {
    try {
      const res = await fetch('/api/info', {
        method: 'POST',
        headers: {'Content-Type':'application/json'},
        body: JSON.stringify({ codes: chunk, token: S.token })
      });
      if (!res.ok) {
        const e = await res.json().catch(() => ({error: res.statusText}));
        showErr(`Ошибка API (${res.status}): ${e.error || res.statusText}`);
        break;
      }
      const data = await res.json();
      S.allResults.push(...data);
      appendRows(data);
    } catch(e) {
      showErr(`Сетевая ошибка: ${e.message}`);
      break;
    }
    done++;
    updateProgress(done, chunks.length);
    if (done < chunks.length) await delay(S.chunkDelay);
  }

  document.getElementById('progress-wrap').classList.add('hidden');
  document.getElementById('btn-check').disabled = false;
  document.getElementById('btn-csv').disabled = false;
  document.getElementById('btn-json').disabled = false;
  updateSummary();
  updateResultLabel();
}

function delay(ms) { return new Promise(r => setTimeout(r, ms)); }

// ── Table Rendering ────────────────────────────────────────────────────────
// New API response item fields:
//   code, status, extendedStatus, packageType, gtin,
//   expirationDate, emissionDate,
//   issuerShortInfo: { issuerTin, issuerName:{ru,uz,en} }

function appendRows(items) {
  const tbody = document.getElementById('result-body');
  items.forEach(item => {
    const cis = item.code || '';
    const isNotFound = !item.status; // API returns minimal object if code not found
    const sc = STATUS[item.status] || { label: item.status || 'Не найден', bg:'#1c0a0a', color:'#9ca3af', border:'#374151' };
    const expDate = item.expirationDate ? new Date(item.expirationDate).toLocaleDateString('ru-RU') : '—';
    const expired = item.expirationDate && new Date(item.expirationDate) < new Date();
    const issuerTin = item.issuerShortInfo?.issuerTin || '—';
    const issuerName = item.issuerShortInfo?.issuerName;
    const issuerNameStr = issuerName
      ? (issuerName.ru || issuerName.uz || issuerName.en || '')
      : '';
    const pkgType = item.packageType || '—';
    const gtin = item.gtin || '—';
    const statusKey = item.status || 'NOT_FOUND';

    const tr = document.createElement('tr');
    tr.className = 'result-row row-enter';
    tr.dataset.cis = cis;
    tr.dataset.status = statusKey;
    tr.style.cursor = 'pointer';

    tr.innerHTML = `
      <td class="px-3 py-2.5">
        <input type="checkbox" class="row-chk" data-cis="${esc(cis)}" onchange="onChkChange(this)">
      </td>
      <td class="px-3 py-2.5 mono text-xs text-gray-300 max-w-[200px] truncate copy-cell"
          title="${esc(cis)}" onclick="copyCell(this)">${esc(cis)}</td>
      <td class="px-3 py-2.5 mono text-xs text-gray-500 copy-cell"
          onclick="copyCell(this)" title="${esc(gtin)}">${esc(gtin)}</td>
      <td class="px-3 py-2.5 text-xs text-gray-500">${esc(pkgType)}</td>
      <td class="px-3 py-2.5 text-sm ${expired?'text-red-400':'text-gray-400'}">${expDate}</td>
      <td class="px-3 py-2.5 mono text-xs text-gray-400 copy-cell"
          onclick="copyCell(this)" title="${esc(issuerTin)}">${esc(issuerTin)}</td>
      <td class="px-3 py-2.5">
        <span class="status-badge" style="background:${sc.bg};color:${sc.color};border-color:${sc.border}">
          ${sc.label}
        </span>
      </td>
      <td class="px-3 py-2.5">
        ${item.extendedStatus
          ? `<span class="status-badge" style="background:#111;color:#6b7280;border-color:#1f2937">${esc(item.extendedStatus)}</span>`
          : ''}
      </td>`;

    tr.addEventListener('click', e => {
      if (e.target.type==='checkbox' || e.target.classList.contains('copy-cell')) return;
      toggleExpand(tr, item, issuerNameStr);
    });
    tbody.appendChild(tr);

    // Expandable detail row
    const exp = document.createElement('tr');
    exp.className = 'expand-row hidden';
    exp.id = `exp-${safeId(cis)}`;
    tbody.appendChild(exp);
  });
}

// ── Expandable Detail Row ──────────────────────────────────────────────────
function safeId(cis) {
  return cis.replace(/[^a-zA-Z0-9]/g, '').substring(0, 16);
}

function toggleExpand(tr, item, issuerNameStr) {
  const cis = item.code || '';
  const expRow = document.getElementById(`exp-${safeId(cis)}`);
  if (!expRow) return;
  const isOpen = !expRow.classList.contains('hidden');
  if (isOpen) {
    expRow.classList.add('hidden');
    tr.classList.remove('expanded-open');
    return;
  }
  tr.classList.add('expanded-open');
  expRow.innerHTML = `<td colspan="8" class="px-4 py-3">
    <div class="text-xs text-gray-600">Загрузка истории...</div></td>`;
  expRow.classList.remove('hidden');
  loadDetail(cis, item, issuerNameStr, expRow);
}

async function loadDetail(cis, item, issuerNameStr, expRow) {
  // Static fields from publicInfo response
  const fields = [
    ['Код (CIS)',       item.code],
    ['GTIN',           item.gtin],
    ['Статус',         item.status],
    ['Расш. статус',   item.extendedStatus],
    ['Тип упаковки',   item.packageType],
    ['ИНН эмитента',   item.issuerShortInfo?.issuerTin],
    ['Эмитент',        issuerNameStr],
    ['Группа товара',  item.productGroupId],
    ['Дата эмиссии',   item.emissionDate ? new Date(item.emissionDate).toLocaleDateString('ru-RU') : null],
    ['Срок годности',  item.expirationDate ? new Date(item.expirationDate).toLocaleDateString('ru-RU') : null],
    ['Дата выпуска',   item.issueDate ? new Date(item.issueDate).toLocaleDateString('ru-RU') : null],
  ].filter(([,v]) => v !== null && v !== undefined && v !== '' && v !== '—');

  let historyHtml = '';
  try {
    const res = await fetch(`/api/history?code=${encodeURIComponent(cis)}&token=${encodeURIComponent(S.token)}`);
    const events = await res.json();
    if (Array.isArray(events) && events.length) {
      historyHtml = `<div class="mt-3 pt-3 border-t border-gray-900">
        <p class="text-xs text-gray-600 uppercase tracking-widest mb-2">История (${events.length})</p>
        <div class="space-y-0 max-h-40 overflow-y-auto">
          ${events.map(ev => {
            const sc = STATUS[ev.eventChangedCodeStatus] || {color:'#6b7280'};
            const dt = ev.eventDate ? new Date(ev.eventDate).toLocaleString('ru-RU') : '—';
            const evType = EVENT_TYPES[ev.eventType] || ev.eventType || '—';
            return `<div class="flex gap-3 items-center py-1 border-b border-gray-900 text-xs">
              <span class="mono text-gray-600 w-36 shrink-0">${dt}</span>
              <span class="text-gray-400 w-32 shrink-0">${esc(evType)}</span>
              <span class="mono" style="color:${sc.color}">${esc(ev.eventChangedCodeStatus||'')}</span>
              <span class="mono text-gray-600 truncate">${esc(ev.eventSourceId||'')}</span>
              ${ev.receiverTin ? `<span class="mono text-gray-500">→ ${esc(ev.receiverTin)}</span>` : ''}
            </div>`;
          }).join('')}
        </div>
      </div>`;
    }
  } catch(_) {}

  expRow.innerHTML = `<td colspan="8" class="px-4 py-3 border-t border-gray-900/50">
    <div class="grid grid-cols-2 md:grid-cols-3 gap-x-6 gap-y-1 text-xs">
      ${fields.map(([k,v]) => `
        <div class="text-gray-600 py-0.5">${esc(k)}</div>
        <div class="mono text-gray-300 py-0.5 truncate copy-cell"
             onclick="copyCell(this)" title="${esc(String(v))}">${esc(String(v))}</div>`
      ).join('')}
    </div>
    ${historyHtml}
  </td>`;
}

// ── Filter ─────────────────────────────────────────────────────────────────
function setFilter(f) {
  S.activeFilter = f;
  setActiveFilterPill(f);
  document.querySelectorAll('#result-body tr[data-status]').forEach(tr => {
    const expRow = document.getElementById(`exp-${safeId(tr.dataset.cis||'')}`);
    const show = f==='all' || tr.dataset.status===f;
    tr.classList.toggle('hidden', !show);
    if (expRow && !show) expRow.classList.add('hidden');
  });
  updateResultLabel();
}
function setActiveFilterPill(f) {
  // Only affect pills that belong to the check-codes filter (use data-filter attribute).
  // Export-tab pills use data-val and must not be touched here.
  document.querySelectorAll('.filter-pill[data-filter]').forEach(p => p.classList.toggle('active', p.dataset.filter===f));
}

// ── Checkboxes ─────────────────────────────────────────────────────────────
function onChkChange(cb) {
  cb.checked ? S.selectedCis.add(cb.dataset.cis) : S.selectedCis.delete(cb.dataset.cis);
  syncMasterChk();
}
function toggleAll() {
  const master = document.getElementById('chk-all');
  document.querySelectorAll('.row-chk').forEach(cb => {
    if (cb.closest('tr').classList.contains('hidden')) return;
    cb.checked = master.checked;
    master.checked ? S.selectedCis.add(cb.dataset.cis) : S.selectedCis.delete(cb.dataset.cis);
  });
}
function syncMasterChk() {
  const all = [...document.querySelectorAll('.row-chk')].filter(c => !c.closest('tr').classList.contains('hidden'));
  const chk = all.filter(c => c.checked);
  const m = document.getElementById('chk-all');
  m.indeterminate = chk.length > 0 && chk.length < all.length;
  m.checked = all.length > 0 && chk.length === all.length;
}

// ── Export ─────────────────────────────────────────────────────────────────
function getExportData() {
  let data = S.allResults;
  if (S.activeFilter !== 'all') {
    data = data.filter(item => {
      const st = item.status || 'NOT_FOUND';
      return st === S.activeFilter;
    });
  }
  if (S.selectedCis.size > 0) {
    data = data.filter(item => S.selectedCis.has(item.code));
  }
  return data;
}

function exportCSV() {
  const data = getExportData();
  if (!data.length) return;
  const hdr = ['code','status','extendedStatus','packageType','gtin',
               'expirationDate','emissionDate','issuerTin','issuerName'];
  const rows = [hdr.join(',')];
  data.forEach(item => {
    const issuerName = item.issuerShortInfo?.issuerName;
    rows.push([
      item.code, item.status, item.extendedStatus, item.packageType, item.gtin,
      item.expirationDate, item.emissionDate,
      item.issuerShortInfo?.issuerTin,
      issuerName ? (issuerName.ru || issuerName.uz || issuerName.en || '') : ''
    ].map(v => {
      const s = v == null ? '' : String(v);
      return /[,"\n]/.test(s) ? `"${s.replace(/"/g,'""')}"` : s;
    }).join(','));
  });
  download('﻿' + rows.join('\r\n'), 'asl-info.csv', 'text/csv;charset=utf-8;');
}

function exportJSON() {
  const data = getExportData();
  if (!data.length) return;
  download(JSON.stringify(data, null, 2), 'asl-info.json', 'application/json');
}

function download(content, name, mime) {
  const a = Object.assign(document.createElement('a'), {
    href: URL.createObjectURL(new Blob([content], {type:mime})),
    download: name
  });
  a.click();
  URL.revokeObjectURL(a.href);
}

// ── Summary ────────────────────────────────────────────────────────────────
function updateSummary() {
  let ok=0, other=0, notFound=0;
  S.allResults.forEach(item => {
    if (!item.status) { notFound++; return; }
    item.status === 'INTRODUCED' ? ok++ : other++;
  });
  document.getElementById('s-total').textContent  = S.allResults.length;
  document.getElementById('s-ok').textContent     = ok;
  document.getElementById('s-other').textContent  = other;
  document.getElementById('s-err').textContent    = notFound;
}
function resetSummary() {
  ['s-total','s-ok','s-other','s-err'].forEach(id => document.getElementById(id).textContent = '0');
}

// ── Progress ───────────────────────────────────────────────────────────────
function updateProgress(done, total) {
  const pct = total ? Math.round(done/total*100) : 0;
  document.getElementById('progress-bar').style.width = `${pct}%`;
  const rem = total - done;
  const etaSec = Math.ceil(rem * S.chunkDelay / 1000);
  const eta = done < total && etaSec > 1 ? ` · ~${etaSec} сек` : '';
  document.getElementById('progress-text').textContent = `${done} / ${total} чанков${eta}`;
}

// ── Error ──────────────────────────────────────────────────────────────────
function showErr(msg) {
  document.getElementById('err-text').textContent = msg;
  document.getElementById('err-banner').classList.remove('hidden');
}
function hideErr() { document.getElementById('err-banner').classList.add('hidden'); }

// ── Reset ──────────────────────────────────────────────────────────────────
function resetInfo() {
  document.getElementById('code-input').value = '';
  document.getElementById('code-count').textContent = '0 кодов';
  S.allResults = []; S.selectedCis = new Set(); S.activeFilter = 'all';
  clearTable();
  hideErr();
  ['summary','filter-wrap','result-wrap','progress-wrap'].forEach(id =>
    document.getElementById(id).classList.add('hidden'));
  document.getElementById('btn-csv').disabled = true;
  document.getElementById('btn-json').disabled = true;
}
function clearTable() {
  document.getElementById('result-body').innerHTML = '';
  const m = document.getElementById('chk-all');
  m.checked = false; m.indeterminate = false;
}
function updateResultLabel() {
  const visible = document.querySelectorAll('#result-body tr[data-status]:not(.hidden)').length;
  document.getElementById('result-label').textContent = `Показано: ${visible} из ${S.allResults.length}`;
}

// ── Copy ───────────────────────────────────────────────────────────────────
function copyCell(el) {
  const t = el.title || el.textContent.trim();
  navigator.clipboard?.writeText(t).then(() => {
    el.classList.add('flash');
    setTimeout(() => el.classList.remove('flash'), 500);
  });
}

// ── History Panel ──────────────────────────────────────────────────────────
async function runHistory() {
  const code = document.getElementById('history-input').value.trim();
  if (!code) return;
  if (!S.token) { alert('Введите токен.'); return; }

  document.getElementById('history-loading').classList.remove('hidden');
  document.getElementById('history-results').innerHTML = '';

  try {
    const res = await fetch(`/api/history?code=${encodeURIComponent(code)}&token=${encodeURIComponent(S.token)}`);
    const data = await res.json();
    if (!res.ok) {
      document.getElementById('history-results').innerHTML =
        `<div class="card p-4 text-red-400 text-sm">${esc(data.error || 'Ошибка')}</div>`;
      return;
    }
    renderTimeline(Array.isArray(data) ? data : []);
  } catch(e) {
    document.getElementById('history-results').innerHTML =
      `<div class="card p-4 text-red-400 text-sm">Сетевая ошибка: ${esc(e.message)}</div>`;
  } finally {
    document.getElementById('history-loading').classList.add('hidden');
  }
}

function renderTimeline(events) {
  const c = document.getElementById('history-results');
  if (!events.length) {
    c.innerHTML = '<div class="card p-6 text-center text-gray-600 text-sm">История не найдена</div>';
    return;
  }
  c.innerHTML = `<p class="text-xs text-gray-600 mono mb-3">${events.length} событий</p>` +
    events.map((ev, idx) => {
      const sc = STATUS[ev.eventChangedCodeStatus] || { color:'#6b7280', bg:'#111', border:'#1f2937' };
      const dt = ev.eventDate ? new Date(ev.eventDate).toLocaleString('ru-RU') : '—';
      const evType = EVENT_TYPES[ev.eventType] || ev.eventType || '—';
      const isLast = idx === events.length - 1;
      return `
      <div class="flex gap-3">
        <div class="flex flex-col items-center" style="width:10px">
          <div class="timeline-dot"></div>
          ${isLast ? '' : '<div class="timeline-line"></div>'}
        </div>
        <div class="flex-1 card p-4 mb-3">
          <div class="flex items-start justify-between gap-2 mb-2">
            <span class="mono text-xs text-gray-600">${dt}</span>
            <span class="status-badge" style="background:${sc.bg||'#111'};color:${sc.color};border-color:${sc.border||'#1f2937'}">
              ${esc(evType)}
            </span>
          </div>
          <div class="grid grid-cols-2 gap-x-4 gap-y-1 text-xs">
            ${ev.eventChangedCodeStatus ? `
              <div class="text-gray-600">Статус после</div>
              <div class="mono" style="color:${sc.color}">${esc(ev.eventChangedCodeStatus)}</div>` : ''}
            ${ev.senderTin ? `
              <div class="text-gray-600">Отправитель</div>
              <div class="mono text-gray-300 copy-cell" onclick="copyCell(this)"
                   title="${esc(ev.senderTin)}">${esc(ev.senderTin)}</div>` : ''}
            ${ev.receiverTin ? `
              <div class="text-gray-600">Получатель</div>
              <div class="mono text-gray-300 copy-cell" onclick="copyCell(this)"
                   title="${esc(ev.receiverTin)}">${esc(ev.receiverTin)}</div>` : ''}
            ${ev.eventSourceId ? `
              <div class="text-gray-600">Документ</div>
              <div class="mono text-gray-500 truncate copy-cell" onclick="copyCell(this)"
                   title="${esc(ev.eventSourceId)}">${esc(ev.eventSourceId)}</div>` : ''}
          </div>
        </div>
      </div>`;
    }).join('');
}

// ════════════════════════════════════════════════════════════════════════════
// ── DOCS TAB ─────────────────────────────────────────────────────────────
// ════════════════════════════════════════════════════════════════════════════

const STATUS_DOC = {
  SUCCESS:             { label:'Успешно',       bg:'#052e1a', color:'#22c55e', border:'#14532d' },
  PARTIALLY_PROCESSED: { label:'Частично',       bg:'#1c1400', color:'#fbbf24', border:'#451a03' },
  ERROR:               { label:'Ошибка',         bg:'#1c0a0a', color:'#f87171', border:'#450a0a' },
  IN_PROCESS:          { label:'В обработке',    bg:'#0c1a2e', color:'#60a5fa', border:'#1e3a5f' },
  IN_PROCESSING:       { label:'Обрабатывается', bg:'#0c1a2e', color:'#60a5fa', border:'#1e3a5f' },
  VALIDATING:          { label:'Валидация',      bg:'#0c1a2e', color:'#93c5fd', border:'#1e3a5f' },
  CREATED:             { label:'Создан',         bg:'#111',    color:'#6b7280', border:'#1f2937' },
  EXPIRED:             { label:'Просрочен',      bg:'#1c0a00', color:'#fb923c', border:'#431407' },
};

const DOC_TYPES = {
  ORIGINAL_INVOICE:              'ЭСФ (стандартная)',
  ADDITIONAL_INVOICE:            'ЭСФ (дополнительная)',
  CORRECTED_INVOICE:             'ЭСФ (исправленная)',
  CANCELLED_INVOICE:             'ЭСФ (аннулированная)',
  SALES_RECEIPT:                 'Чек ККТ (продажа)',
  REFUND_RECEIPT:                'Чек ККТ (возврат)',
  ORDER:                         'Заказ КМ',
  ORIGINAL_CUSTOMS_DECLARATION:  'Таможенная декларация',
  CUSTOMS_CODE_WITHDRAWAL:       'Вывод через таможню',
  CUSTOMS_CODE_REGISTRATION:     'Регистрация АИК',
  UTILISATION:                   'Нанесение КМ',
  AGGREGATION:                   'Агрегация',
  TRANSPORT_CODE_DISAGGREGATION: 'Расформирование',
  TRANSFER_REQUEST:              'Передача КМ',
  WRITE_OFF_NOTICE:              'Списание',
  VALIDATION:                    'Валидация печати',
  WITHDRAWAL:                    'Вывод из оборота',
};

const CODE_STATE = {
  SUCCESS: { icon:'✅', color:'#22c55e' },
  WARNING:  { icon:'⚠️', color:'#fbbf24' },
  ERROR:    { icon:'❌', color:'#f87171' },
};

// Состояния кодов маркировки в чеках ККТ
const RCP_MARK_STATE = {
  INTRODUCED:  { label:'В обороте',            color:'#22c55e' },
  WITHDRAWN:   { label:'Выведен из оборота',   color:'#a78bfa' },
  APPLIED:     { label:'Нанесён',              color:'#34d399' },
  RECEIVED:    { label:'Получен',              color:'#60a5fa' },
  WRITTEN_OFF: { label:'Списан',               color:'#9ca3af' },
  CHANGED:     { label:'Изменено',             color:'#fbbf24' },
  RETIRED:     { label:'Аннулирован',          color:'#f87171' },
  EMITTED:     { label:'Эмитирован',           color:'#c4f542' },
  SUCCESS:     { label:'Успешно',              color:'#22c55e' },
  ERROR:       { label:'Ошибка',               color:'#f87171' },
  WARNING:     { label:'Предупреждение',       color:'#fbbf24' },
  SOLD:        { label:'Продан',               color:'#a78bfa' },
  RETURNED:    { label:'Возвращён',            color:'#60a5fa' },
  CREATED:     { label:'Создан',               color:'#9ca3af' },
  EXPORTED:    { label:'Экспортирован',        color:'#34d399' },
};

function toggleDocFilter(type, val, btn) {
  const set = type === 'status' ? S.docsStatusFilter : S.docsTypesFilter;
  set.has(val) ? set.delete(val) : set.add(val);
  btn.classList.toggle('active', set.has(val));
  if (!document.getElementById('docs-result-wrap').classList.contains('hidden')) {
    loadDocs();
  }
}

async function loadDocById() {
  const docId = document.getElementById('doc-direct-id').value.trim();
  if (!docId) return;
  if (!S.token) { showDocsErr('Введите токен.'); return; }
  document.getElementById('docs-err').classList.add('hidden');
  document.getElementById('docs-loading').classList.remove('hidden');
  try {
    const res = await fetch(`/api/docs/${encodeURIComponent(docId)}?token=${encodeURIComponent(S.token)}`);
    const data = await res.json();
    if (!res.ok) { showDocsErr(data.error || res.statusText); return; }
    // Build a minimal doc object and show it in the table
    const doc = Array.isArray(data) ? data[0] : (data.documentInfos?.[0] ?? data);
    if (!doc || !doc.documentId) { showDocsErr('Документ не найден или неверный ID.'); return; }
    S.docs = [doc];
    document.getElementById('docs-result-wrap').classList.remove('hidden');
    document.getElementById('docs-tbody').innerHTML = '';
    appendDocRows([doc]);
    document.getElementById('docs-count-label').textContent = 'Найдено: 1';
    document.getElementById('docs-load-more').classList.add('hidden');
    // Auto-expand the card
    const tr = document.querySelector(`tr[data-doc-id="${CSS.escape(doc.documentId)}"]`);
    if (tr) tr.click();
  } catch(e) {
    showDocsErr('Сетевая ошибка: ' + e.message);
  } finally {
    document.getElementById('docs-loading').classList.add('hidden');
  }
}

function toggleExpFilter(type, val, btn) {
  const set = type === 'status' ? S.expStatusFilter : S.expPkgFilter;
  set.has(val) ? set.delete(val) : set.add(val);
  btn.classList.toggle('active', set.has(val));
}

let _docsReqId = 0;

async function loadDocs() {
  if (!S.token) { showDocsErr('Введите токен.'); return; }
  const myReqId = ++_docsReqId;

  S.docs = [];
  document.getElementById('docs-tbody').innerHTML = '';
  document.getElementById('docs-result-wrap').classList.remove('hidden');
  document.getElementById('docs-load-more').classList.add('hidden');
  document.getElementById('docs-loading').classList.remove('hidden');
  document.getElementById('docs-err').classList.add('hidden');
  document.getElementById('docs-count-label').textContent = 'Загрузка…';

  const df = document.getElementById('doc-date-from').value;
  const dt = document.getElementById('doc-date-to').value;

  let cursor = null;
  let pageNum = 0;

  try {
    while (true) {
      if (myReqId !== _docsReqId) return;

      const params = new URLSearchParams({ token: S.token, limit: 50 });
      S.docsStatusFilter.forEach(v => params.append('status', v));
      S.docsTypesFilter.forEach(v => params.append('types', v));
      if (df) params.append('dateFrom', df + 'T00:00:00Z');
      if (dt) params.append('dateTo',   dt + 'T23:59:59Z');
      if (cursor) params.append('cursor', cursor);

      const res = await fetch('/api/docs?' + params);
      if (myReqId !== _docsReqId) return;
      const data = await res.json();
      if (myReqId !== _docsReqId) return;
      if (!res.ok) { showDocsErr(data.error || res.statusText); return; }

      const rawItems = Array.isArray(data) ? data : (data.documentInfos || []);

      // Client-side type guard
      const typed = S.docsTypesFilter.size > 0
        ? rawItems.filter(d => S.docsTypesFilter.has(d.type))
        : rawItems;

      // Deduplicate
      const seenIds = new Set(S.docs.map(d => d.documentId));
      const items = typed.filter(d => !seenIds.has(d.documentId));

      S.docs.push(...items);
      if (items.length > 0) appendDocRows(items);

      document.getElementById('docs-count-label').textContent =
        rawItems.length < 50
          ? `Найдено: ${S.docs.length}`
          : `Загружено: ${S.docs.length}…`;

      // Next cursor: prefer API-returned cursor, fallback to last documentId
      const nextCursor = (data.cursor || data.nextCursor || data.nextId ||
        (rawItems.length >= 50 ? rawItems[rawItems.length - 1].documentId : null));

      if (!nextCursor || rawItems.length < 50) break;
      cursor = nextCursor;
      pageNum++;

      // Show "loading next" indicator
      document.getElementById('docs-load-more').classList.remove('hidden');
    }

    document.getElementById('docs-load-more').classList.add('hidden');
    document.getElementById('docs-count-label').textContent = `Найдено: ${S.docs.length}`;
    document.getElementById('btn-docs-excel').disabled = S.docs.length === 0;
    document.getElementById('btn-bulk-codes').disabled = S.docs.length === 0;

    // Fetch FacturaNo/FacturaId from doc details in background
    if (S.docs.length > 0 && myReqId === _docsReqId) enrichFacturaNos(myReqId);
  } catch(e) {
    if (myReqId !== _docsReqId) return;
    showDocsErr('Сетевая ошибка: ' + e.message);
  } finally {
    if (myReqId === _docsReqId) document.getElementById('docs-loading').classList.add('hidden');
  }
}

// ── Extract FacturaNo from any known response shape ──────────────────────────
function _extractFacturaFields(obj) {
  if (!obj || typeof obj !== 'object') return null;
  // Unwrap arrays/wrappers
  const d = Array.isArray(obj) ? obj[0]
          : (obj.documentInfos?.[0] || obj.document || obj.invoice || obj.body || obj.data || obj);
  if (!d) return null;

  // Try every known field name for invoice number
  const noKeys = ['facturaNo','FacturaNo','invoiceNo','invoiceNumber','docNo',
                  'documentNo','number','contractNo','formattedNumber'];
  const idKeys = ['facturaId','FacturaId','invoiceId','docId','externalId'];

  let no = '', id = '';
  for (const k of noKeys) { if (d[k] && String(d[k]).trim()) { no = String(d[k]).trim(); break; } }
  for (const k of idKeys) { if (d[k] && String(d[k]).trim()) { id = String(d[k]).trim(); break; } }

  // Recurse one level deeper if still not found
  if (!no) {
    for (const sub of ['documentInfos','invoice','body','data','info']) {
      const inner = d[sub];
      if (!inner) continue;
      const r = _extractFacturaFields(Array.isArray(inner) ? inner[0] : inner);
      if (r?.no) return r;
    }
  }
  return no ? { no, id } : null;
}

async function enrichFacturaNos(reqId) {
  const BATCH = 5;
  const label = document.getElementById('docs-count-label');
  const total = S.docs.length;

  for (let i = 0; i < total; i += BATCH) {
    if (reqId !== _docsReqId) return; // cancelled by new search
    const batch = S.docs.slice(i, i + BATCH);
    label.textContent = `Загружаю номера ЭСФ: ${Math.min(i + BATCH, total)} / ${total}…`;

    await Promise.all(batch.map(async doc => {
      if (doc._facturaNo) return; // already enriched
      try {
        const res = await fetch(`/api/docs/${encodeURIComponent(doc.documentId)}?token=${encodeURIComponent(S.token)}`);
        if (!res.ok) return;
        const raw = await res.json();
        const f = _extractFacturaFields(raw);
        if (!f) {
          // Debug: store raw keys so user can report
          doc._rawKeys = Object.keys(raw).join(', ');
          return;
        }
        doc._facturaNo = f.no;
        doc._facturaId = f.id || doc.documentId;

        // Update visible row
        const tr = document.querySelector(`tr[data-doc-id="${CSS.escape(doc.documentId)}"]`);
        if (tr) {
          tr.dataset.facturaNo = f.no;
          const c2 = tr.querySelector('td:nth-child(2)');
          const c3 = tr.querySelector('td:nth-child(3)');
          if (c2) { c2.textContent = f.no; c2.title = f.no; }
          if (c3 && f.id) { c3.textContent = f.id; c3.title = f.id; }
        }
      } catch(_) {}
    }));
    await delay(80);
  }
  label.textContent = `Найдено: ${total}`;

  // If FacturaNo still not found — show raw keys for first doc to debug
  const missing = S.docs.filter(d => !d._facturaNo);
  if (missing.length > 0 && missing[0]._rawKeys) {
    document.getElementById('docs-diag').textContent =
      `FacturaNo не найден. Поля API для первого документа: ${missing[0]._rawKeys}`;
    document.getElementById('docs-diag-wrap').classList.remove('hidden');
  }
}

let docsSortAsc = false;
function sortDocsByDate() {
  docsSortAsc = !docsSortAsc;
  S.docs.sort((a, b) => {
    const da = new Date(a.createDate || 0).getTime();
    const db = new Date(b.createDate || 0).getTime();
    return docsSortAsc ? da - db : db - da;
  });
  document.getElementById('docs-tbody').innerHTML = '';
  appendDocRows(S.docs);
  const th = document.getElementById('docs-date-th');
  if (th) th.textContent = 'Дата ' + (docsSortAsc ? '↑' : '↓');
}

function resetDocs() {
  S.docs = []; S.docsCursor = null;
  document.getElementById('docs-tbody').innerHTML = '';
  document.getElementById('docs-count-label').textContent = '';
  document.getElementById('docs-result-wrap').classList.add('hidden');
  document.getElementById('docs-load-more').classList.add('hidden');
  document.getElementById('docs-err').classList.add('hidden');
  document.getElementById('btn-docs-excel').disabled = true;
}

function showDocsErr(msg) {
  document.getElementById('docs-err-text').textContent = msg;
  document.getElementById('docs-err').classList.remove('hidden');
}

function appendDocRows(items) {
  const tbody = document.getElementById('docs-tbody');
  items.forEach(doc => {
    const sc       = STATUS_DOC[doc.status] || { label:doc.status, bg:'#111', color:'#6b7280', border:'#1f2937' };
    const dt       = doc.createDate ? new Date(doc.createDate).toLocaleString('ru-RU') : '—';
    const typeName = DOC_TYPES[doc.type] || doc.type || '—';
    // FacturaNo / FacturaId — могут быть в разных полях в зависимости от версии API
    const facturaNo = doc.facturaNo || doc.FacturaNo || doc.invoiceNumber || doc.originalDocId || '—';
    const facturaId = doc.facturaId || doc.FacturaId || doc.externalId || doc.documentId || '—';

    const tr = document.createElement('tr');
    tr.className = 'result-row row-enter';
    tr.dataset.docId = doc.documentId;
    tr.dataset.facturaNo = facturaNo;
    tr.dataset.facturaId = facturaId;
    tr.style.cursor = 'pointer';
    tr.innerHTML = `
      <td class="px-3 py-2.5"><input type="checkbox" class="doc-chk" data-id="${esc(doc.documentId)}" onchange="syncDocsChk()"></td>
      <td class="px-3 py-2.5 mono text-xs text-gray-300 max-w-[160px] truncate copy-cell"
          onclick="copyCell(this)" title="${esc(facturaNo)}">${esc(facturaNo)}</td>
      <td class="px-3 py-2.5 mono text-xs text-gray-500 max-w-[200px] truncate copy-cell"
          onclick="copyCell(this)" title="${esc(facturaId)}">${esc(facturaId)}</td>
      <td class="px-3 py-2.5 text-xs text-gray-500">${esc(typeName)}</td>
      <td class="px-3 py-2.5">
        <span class="status-badge" style="background:${sc.bg};color:${sc.color};border-color:${sc.border}">${sc.label}</span>
      </td>
      <td class="px-3 py-2.5 text-xs text-gray-500 mono">${dt}</td>`;
    tr.addEventListener('click', e => {
      if (e.target.type==='checkbox'||e.target.classList.contains('copy-cell')) return;
      toggleDocExpand(tr, doc.documentId);
    });
    tbody.appendChild(tr);

    const exp = document.createElement('tr');
    exp.className = 'expand-row hidden';
    exp.id = `docexp-${safeId(doc.documentId)}`;
    tbody.appendChild(exp);
  });
}

const docExpandState = {};

async function toggleDocExpand(tr, docId) {
  const expRow = document.getElementById(`docexp-${safeId(docId)}`);
  if (!expRow) return;
  if (!expRow.classList.contains('hidden')) {
    expRow.classList.add('hidden');
    tr.classList.remove('expanded-open');
    return;
  }
  tr.classList.add('expanded-open');
  expRow.innerHTML = `<td colspan="6" class="px-4 py-3"><div class="text-xs text-gray-600">Загрузка деталей...</div></td>`;
  expRow.classList.remove('hidden');
  await loadDocDetail(docId, expRow);
}

async function loadDocDetail(docId, expRow) {
  const [errRes, codesRes] = await Promise.all([
    fetch(`/api/docs/${encodeURIComponent(docId)}/errors?token=${encodeURIComponent(S.token)}&limit=100`),
    fetch(`/api/docs/${encodeURIComponent(docId)}/codes?token=${encodeURIComponent(S.token)}&limit=200`)
  ]);

  const errData   = errRes.ok   ? (await errRes.json())   : { documentErrors: [] };
  const codesData = codesRes.ok ? (await codesRes.json()) : [];

  const errors = errData.documentErrors || [];
  const codes  = Array.isArray(codesData) ? codesData : (codesData.codes || []);

  // Collect error codes for export
  const errorCodes = codes.filter(c => c.state === 'ERROR' || c.state === 'WARNING').map(c => c.code);

  const errHtml = errors.length ? `
    <div class="mb-3">
      <p class="text-xs text-gray-600 uppercase tracking-widest mb-2">Ошибки документа (${errors.length})</p>
      <div class="overflow-x-auto max-h-40 overflow-y-auto">
        <table class="w-full text-xs border-collapse">
          <tr style="background:#0f0f0f"><th class="text-left px-2 py-1 text-gray-600">#</th><th class="text-left px-2 py-1 text-gray-600">Поле</th><th class="text-left px-2 py-1 text-gray-600">Код ошибки</th><th class="text-left px-2 py-1 text-gray-600">Детали</th></tr>
          ${errors.map((e,i) => `<tr class="border-b border-gray-900">
            <td class="px-2 py-1 text-gray-600">${e.index??i}</td>
            <td class="px-2 py-1 mono text-gray-500">${esc(e.propertyName||'—')}</td>
            <td class="px-2 py-1 mono" style="color:#f87171">${esc(e.errorCode||'—')}</td>
            <td class="px-2 py-1 mono text-gray-600 text-xs truncate max-w-xs">${esc(JSON.stringify(e.errorTags||{}))}</td>
          </tr>`).join('')}
        </table>
      </div>
    </div>` : '';

  const codesHtml = codes.length ? `
    <div>
      <div class="flex items-center justify-between mb-2">
        <p class="text-xs text-gray-600 uppercase tracking-widest">Коды в документе (${codes.length})</p>
        <div class="flex gap-2">
          ${errorCodes.length ? `<button class="btn-ghost text-xs" style="padding:3px 8px"
            onclick="downloadDocCodes(${JSON.stringify(errorCodes)},${JSON.stringify(codes)},'errors-${docId}.xlsx',true)">↓ Проблемные Excel (${errorCodes.length})</button>` : ''}
          <button class="btn-ghost text-xs" style="padding:3px 8px"
            onclick="downloadDocCodes(${JSON.stringify(codes.map(c=>c.code))},${JSON.stringify(codes)},'all-${docId}.xlsx',false)">↓ Все Excel</button>
        </div>
      </div>
      <div class="max-h-48 overflow-y-auto">
        <table class="w-full text-xs border-collapse">
          <tr style="background:#0f0f0f"><th class="text-left px-2 py-1 text-gray-600">#</th><th class="text-left px-2 py-1 text-gray-600">Код</th><th class="text-left px-2 py-1 text-gray-600">Состояние</th><th class="text-left px-2 py-1 text-gray-600">Причина</th></tr>
          ${codes.map(c => {
            const st = CODE_STATE[c.state] || { icon:'·', color:'#6b7280' };
            return `<tr class="border-b border-gray-900">
              <td class="px-2 py-1 text-gray-600">${c.index??''}</td>
              <td class="px-2 py-1 mono text-gray-300 copy-cell" onclick="copyCell(this)" title="${esc(c.code)}">${esc(c.code||'—')}</td>
              <td class="px-2 py-1" style="color:${st.color}">${st.icon} ${esc(c.state||'—')}</td>
              <td class="px-2 py-1 mono text-gray-600">${esc(c.result||'')}</td>
            </tr>`;
          }).join('')}
        </table>
      </div>
    </div>` : '<p class="text-xs text-gray-600">Коды не найдены</p>';

  expRow.innerHTML = `<td colspan="6" class="px-4 py-3 border-t border-gray-900/50">
    ${errHtml}${codesHtml}
  </td>`;
}

async function downloadDocCodes(filteredCodes, allCodeObjs, filename, errorsOnly) {
  const rows = errorsOnly
    ? allCodeObjs.filter(c => filteredCodes.includes(c.code))
    : allCodeObjs;
  try {
    const res = await fetch('/api/doc-codes-excel', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({ codes: rows, filename })
    });
    if (!res.ok) { alert('Ошибка генерации Excel: ' + res.statusText); return; }
    const blob = await res.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url; a.download = filename; a.click();
    URL.revokeObjectURL(url);
  } catch(e) { alert('Ошибка: ' + e.message); }
}

function toggleDocsAll() {
  const m = document.getElementById('docs-chk-all');
  document.querySelectorAll('.doc-chk').forEach(c => { c.checked = m.checked; });
}
function syncDocsChk() {
  const all = document.querySelectorAll('.doc-chk');
  const chk = document.querySelectorAll('.doc-chk:checked');
  const m = document.getElementById('docs-chk-all');
  m.indeterminate = chk.length > 0 && chk.length < all.length;
  m.checked = all.length > 0 && chk.length === all.length;
}

async function exportDocsExcel() {
  if (!S.docs.length) return;
  try {
    const res = await fetch('/api/docs-list-excel', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({ docs: S.docs })
    });
    if (!res.ok) { alert('Ошибка генерации Excel: ' + res.statusText); return; }
    const blob = await res.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url; a.download = 'dokumenty.xlsx'; a.click();
    URL.revokeObjectURL(url);
  } catch(e) { alert('Ошибка: ' + e.message); }
}

async function downloadAllDocCodes() {
  if (!S.docs.length) return;
  const prog     = document.getElementById('bulk-progress-wrap');
  const progText = document.getElementById('bulk-progress-text');
  prog.classList.remove('hidden');
  document.getElementById('btn-bulk-codes').disabled = true;

  const docsWithCodes = [];
  let totalCodes = 0;

  for (let i = 0; i < S.docs.length; i++) {
    const doc = S.docs[i];
    if (i % 5 === 0) await delay(100); // лёгкая задержка каждые 5 запросов
    const facturaNo = doc._facturaNo || doc.facturaNo || doc.FacturaNo || doc.originalDocId || doc.documentId;
    const facturaId = doc._facturaId || doc.facturaId || doc.FacturaId || doc.documentId;

    progText.textContent = `Загружаю коды: документ ${i + 1} из ${S.docs.length} (${facturaNo})…`;

    try {
      // Load ALL code pages for this document
      let allCodes = [];
      let lastIndex = 0;
      while (true) {
        const params = new URLSearchParams({ token: S.token, limit: 200 });
        if (lastIndex) params.append('lastIndex', lastIndex);
        const res = await fetch(`/api/docs/${encodeURIComponent(doc.documentId)}/codes?` + params);
        if (!res.ok) break;
        const data = await res.json();
        const page = Array.isArray(data) ? data : (data.codes || []);
        allCodes.push(...page);
        if (page.length < 200) break;
        lastIndex = allCodes.length;
      }
      if (allCodes.length > 0) {
        docsWithCodes.push({ facturaNo, facturaId, docId: doc.documentId, type: doc.type, codes: allCodes });
        totalCodes += allCodes.length;
      }
    } catch(_) {}
  }

  prog.classList.add('hidden');
  document.getElementById('btn-bulk-codes').disabled = false;

  if (!totalCodes) { alert('Коды маркировки не найдены ни в одном документе.'); return; }

  progText.textContent = `Генерирую Excel (${totalCodes} кодов)…`;
  prog.classList.remove('hidden');

  try {
    const res = await fetch('/api/docs-bulk-codes-excel', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({ docs: docsWithCodes })
    });
    if (!res.ok) { alert('Ошибка генерации Excel: ' + res.statusText); return; }
    const blob = await res.blob();
    const url  = URL.createObjectURL(blob);
    const a    = document.createElement('a');
    a.href = url; a.download = 'vse-kody-esf.xlsx'; a.click();
    URL.revokeObjectURL(url);
  } catch(e) {
    alert('Ошибка: ' + e.message);
  } finally {
    prog.classList.add('hidden');
  }
}

// ════════════════════════════════════════════════════════════════════════════
// ── RECEIPTS (ЧЕК ККТ) ───────────────────────────────────────────────────
// ════════════════════════════════════════════════════════════════════════════

const RCP_TYPE_RU = { SALE: 'Продажа', RETURN: 'Возврат', SALES: 'Продажа', REFUND: 'Возврат' };
const RCP_STATUS = {
  SUCCESS:             { label:'Обработан успешно', bg:'#052e1a', color:'#22c55e', border:'#14532d' },
  ERROR:               { label:'Ошибка',            bg:'#1c0a0a', color:'#f87171', border:'#450a0a' },
  PARTIALLY_PROCESSED: { label:'Частично',          bg:'#1c1400', color:'#fbbf24', border:'#451a03' },
  IN_PROCESS:          { label:'В обработке',       bg:'#0c1a2e', color:'#60a5fa', border:'#1e3a5f' },
  CREATED:             { label:'Создан',            bg:'#111',    color:'#6b7280', border:'#1f2937' },
};

let rcpTypeFilter = new Set(['SALE','RETURN']);
let rcpAllData = [];

async function syncRcpWithToken() {
  const raw = document.getElementById('rcp-session-token').value.trim();
  const token = raw.replace(/^bearer\s+/i, '');
  if (!token) { alert('Вставьте токен из заголовка Authorization'); return; }

  const prog = document.getElementById('rcp-sync-progress');
  const ok   = document.getElementById('rcp-sync-ok');
  prog.classList.remove('hidden');
  ok.classList.add('hidden');

  try {
    const res = await fetch('/api/rcp-sync', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({ token })
    });
    const data = await res.json();
    if (!res.ok || data.error) {
      alert('Ошибка синхронизации: ' + (data.error || res.statusText));
      return;
    }
    ok.textContent = `Синхронизировано чеков: ${data.count}  ·  ${(data.synced_at||'').replace('T',' ')}`;
    ok.classList.remove('hidden');
    document.getElementById('rcp-sync-status').textContent = `Всего в кеше: ${data.count}`;
    document.getElementById('rcp-sync-time').textContent = 'Синхр.: ' + (data.synced_at||'').replace('T',' ');
    // Auto-load receipts after sync
    applyRcpFilter();
  } catch(e) {
    alert('Сетевая ошибка: ' + e.message);
  } finally {
    prog.classList.add('hidden');
  }
}

function switchDocsSubtab(name) {
  ['doclist','receipts'].forEach(t => {
    document.getElementById(`subtab-btn-${t}`).classList.toggle('subtab-active', t === name);
    document.getElementById(`subpanel-${t}`).classList.toggle('hidden', t !== name);
  });
  if (name === 'receipts') refreshRcpSyncStatus();
}

async function refreshRcpSyncStatus() {
  try {
    const res = await fetch('/api/rcp-cached');
    const data = await res.json();
    if (data.synced_at) {
      document.getElementById('rcp-sync-status').textContent = `Всего в кеше: ${data.cached_total}`;
      document.getElementById('rcp-sync-time').textContent = 'Синхр.: ' + data.synced_at.replace('T',' ');
    }
  } catch(e) {}
}

function toggleRcpType(val, btn) {
  rcpTypeFilter.has(val) ? rcpTypeFilter.delete(val) : rcpTypeFilter.add(val);
  btn.classList.toggle('active', rcpTypeFilter.has(val));
  applyRcpFilter();
}

async function applyRcpFilter() {
  document.getElementById('rcp-err').classList.add('hidden');
  document.getElementById('rcp-loading').classList.remove('hidden');

  const df  = document.getElementById('rcp-date-from').value;
  const dt  = document.getElementById('rcp-date-to').value;
  const ext = document.getElementById('rcp-ext-id-filter').value.trim();

  try {
    const params = new URLSearchParams();
    if (df) params.append('dateFrom', df);
    if (dt) params.append('dateTo', dt);
    if (ext) params.append('externalId', ext);
    const res = await fetch('/api/rcp-cached?' + params);
    const data = await res.json();

    if (!data.synced_at) {
      showRcpErr('Нет данных. Войдите в ЛК ASL Belgisi, выполните скрипт в консоли, затем нажмите «Загрузить».');
      document.getElementById('rcp-result-wrap').classList.add('hidden');
      return;
    }

    // Update sync status
    document.getElementById('rcp-sync-status').textContent = `Всего в кеше: ${data.cached_total}`;
    document.getElementById('rcp-sync-time').textContent = 'Синхр.: ' + (data.synced_at || '').replace('T',' ');

    let items = data.receipts || [];
    // Apply type filter client-side
    if (rcpTypeFilter.size < 2) {
      items = items.filter(r => {
        const t = (r.type || '').toUpperCase().replace('SALES','SALE').replace('REFUND','RETURN');
        return rcpTypeFilter.has(t);
      });
    }

    rcpAllData = items;
    document.getElementById('rcp-result-wrap').classList.remove('hidden');
    document.getElementById('rcp-tbody').innerHTML = '';
    appendRcpRows(items);
    document.getElementById('rcp-count-label').textContent = `Показано: ${items.length}`;
    document.getElementById('btn-rcp-excel').disabled = items.length === 0;
    document.getElementById('btn-rcp-bulk').disabled = items.length === 0;
  } catch(e) {
    showRcpErr('Ошибка: ' + e.message);
  } finally {
    document.getElementById('rcp-loading').classList.add('hidden');
  }
}

function resetRcpFilter() {
  document.getElementById('rcp-date-from').value = '';
  document.getElementById('rcp-date-to').value = '';
  document.getElementById('rcp-ext-id-filter').value = '';
  rcpTypeFilter = new Set(['SALE','RETURN']);
  document.querySelectorAll('[data-rcp-type]').forEach(b => b.classList.add('active'));
  applyRcpFilter();
}

function appendRcpRows(items) {
  const tbody = document.getElementById('rcp-tbody');
  items.forEach(r => {
    const sc = RCP_STATUS[r.status] || { label: r.status || '—', bg:'#111', color:'#6b7280', border:'#1f2937' };
    const seller = r.seller || r.sellerInfo || {};
    const sellerTin  = seller.tin || seller.sellerTin || r.sellerTin || '—';
    const sellerName = (() => {
      const n = seller.name || seller.sellerName || r.sellerName || '';
      return typeof n === 'object' ? (n.ru || n.uz || '') : n;
    })();
    const typeRu   = RCP_TYPE_RU[(r.type||'').toUpperCase()] || r.type || '—';
    const created  = r.createdOn || r.createdDate || r.createDate || '';
    const createdFmt = created ? new Date(created).toLocaleString('ru-RU') : '—';
    const extId    = r.externalId || r.originalDocId || '—';

    const tr = document.createElement('tr');
    tr.className = 'result-row row-enter';
    tr.dataset.rcpId = r.id;
    tr.style.cursor = 'pointer';
    tr.innerHTML = `
      <td class="px-3 py-2.5 mono text-xs text-gray-300 max-w-[220px] truncate copy-cell"
          onclick="copyCell(this)" title="${esc(r.id||'')}">${esc(r.id||'—')}</td>
      <td class="px-3 py-2.5 mono text-xs text-gray-600 max-w-[200px] truncate copy-cell"
          onclick="copyCell(this)" title="${esc(extId)}">${esc(extId)}</td>
      <td class="px-3 py-2.5 text-xs text-gray-400">${esc(typeRu)}</td>
      <td class="px-3 py-2.5">
        <span class="status-badge" style="background:${sc.bg};color:${sc.color};border-color:${sc.border}">${sc.label}</span>
      </td>
      <td class="px-3 py-2.5 text-xs text-gray-500 mono">${createdFmt}</td>
      <td class="px-3 py-2.5 text-xs text-gray-500">${esc(sellerTin)} ${sellerName ? '<span class="text-gray-700">· '+esc(sellerName)+'</span>' : ''}</td>`;
    tr.addEventListener('click', e => {
      if (e.target.classList.contains('copy-cell')) return;
      toggleRcpExpand(tr, r);
    });
    tbody.appendChild(tr);

    const exp = document.createElement('tr');
    exp.className = 'expand-row hidden';
    exp.id = `rcpexp-${safeId(r.id||'')}`;
    tbody.appendChild(exp);
  });
}

async function toggleRcpExpand(tr, r) {
  const expRow = document.getElementById(`rcpexp-${safeId(r.id||'')}`);
  if (!expRow) return;
  if (!expRow.classList.contains('hidden')) {
    expRow.classList.add('hidden');
    tr.classList.remove('expanded-open');
    return;
  }
  tr.classList.add('expanded-open');
  expRow.innerHTML = `<td colspan="6" class="px-4 py-3"><div class="flex items-center gap-2 text-xs text-gray-600"><div class="spinner" style="width:14px;height:14px;border-width:2px"></div> Загружаю детали чека...</div></td>`;
  expRow.classList.remove('hidden');

  try {
    const res = await fetch(`/api/rcp-detail/${encodeURIComponent(r.id)}`);
    const detail = await res.json();
    if (!res.ok) {
      expRow.innerHTML = `<td colspan="6" class="px-4 py-3 text-xs text-red-400">${esc(detail.error || 'Ошибка загрузки')}</td>`;
      return;
    }
    renderRcpDetail(expRow, detail, r);
  } catch(e) {
    expRow.innerHTML = `<td colspan="6" class="px-4 py-3 text-xs text-red-400">Сетевая ошибка: ${esc(e.message)}</td>`;
  }
}

function renderRcpDetail(expRow, detail, r) {
  // API returns the receipt object directly or wrapped
  const d = detail.receipt || detail;

  // Extract marking codes — try all known field names
  let codes = [];
  for (const key of ['markingCodes', 'codes', 'cises', 'markCodes', 'items']) {
    if (Array.isArray(d[key]) && d[key].length > 0) { codes = d[key]; break; }
  }

  // Seller info
  const seller = d.seller || d.sellerInfo || r.seller || {};
  const sellerTin  = seller.tin || seller.sellerTin || d.sellerTin || r.sellerTin || '—';
  const sellerName = (() => { const n = seller.name || seller.sellerName || ''; return typeof n === 'object' ? (n.ru || n.uz || '') : n; })();

  const fields = [
    ['ID чека',           d.id || r.id],
    ['Внешний ID',        d.externalId || r.externalId],
    ['Тип',              RCP_TYPE_RU[(d.type||r.type||'').toUpperCase()] || d.type || r.type],
    ['Статус',           (RCP_STATUS[d.status||r.status] || {label: d.status||r.status||'—'}).label],
    ['Дата регистрации', (d.createdOn || d.createdDate || r.createdOn || '').replace('T',' ').slice(0,19)],
    ['Дата чека',        (d.receiptDate || d.dateTime || d.checkDate || '').replace('T',' ').slice(0,19)],
    ['ИНН продавца',     sellerTin],
    ['Продавец',         sellerName],
    ['Источник',         d.source || d.sourceType || ''],
    ['Адрес продажи',    d.salesAddress || d.address || ''],
  ].filter(([,v]) => v && v !== '—');

  // Позиции чека (товары)
  const positions = d.positions || d.items || d.goods || [];
  const posHtml = positions.length ? `
    <div class="mb-3">
      <p class="text-xs text-gray-600 uppercase tracking-widest mb-2">Позиции чека (${positions.length})</p>
      <div class="overflow-x-auto max-h-40 overflow-y-auto">
        <table class="w-full text-xs border-collapse">
          <tr style="background:#0f0f0f">
            <th class="text-left px-2 py-1 text-gray-600">Наименование</th>
            <th class="text-left px-2 py-1 text-gray-600">Кол-во</th>
            <th class="text-left px-2 py-1 text-gray-600">Цена с НДС</th>
            <th class="text-left px-2 py-1 text-gray-600">Сумма НДС</th>
          </tr>
          ${positions.map(p => `<tr class="border-b border-gray-900">
            <td class="px-2 py-1 text-gray-300">${esc(p.name || p.productName || p.goodName || '—')}</td>
            <td class="px-2 py-1 mono text-gray-500">${esc(String(p.quantity ?? p.count ?? '—'))}</td>
            <td class="px-2 py-1 mono text-gray-500">${esc(String(p.priceWithVat ?? p.price ?? '—'))}</td>
            <td class="px-2 py-1 mono text-gray-500">${esc(String(p.vatSum ?? p.vat ?? '—'))}</td>
          </tr>`).join('')}
        </table>
      </div>
    </div>` : '';

  // Коды маркировки
  const codesHtml = codes.length ? `
    <div>
      <p class="text-xs text-gray-600 uppercase tracking-widest mb-2">Коды маркировки (${codes.length})</p>
      <div class="max-h-52 overflow-y-auto">
        <table class="w-full text-xs border-collapse">
          <tr style="background:#0f0f0f">
            <th class="text-left px-2 py-1 text-gray-600">#</th>
            <th class="text-left px-2 py-1 text-gray-600">Код маркировки</th>
            <th class="text-left px-2 py-1 text-gray-600">GTIN</th>
            <th class="text-left px-2 py-1 text-gray-600">Состояние</th>
          </tr>
          ${codes.map((c, i) => {
            const codeStr = typeof c === 'string' ? c : (c.code || c.cis || c.markCode || c.identificationCode || JSON.stringify(c));
            const gtin    = typeof c === 'object' ? (c.gtin || c.productGtin || '') : '';
            const state   = typeof c === 'object' ? (c.state || c.status || c.markingCodeState || '') : '';
            const stConf  = RCP_MARK_STATE[state] || { label: state || '—', color:'#6b7280' };
            return `<tr class="border-b border-gray-900">
              <td class="px-2 py-1 text-gray-600">${i+1}</td>
              <td class="px-2 py-1 mono text-gray-300 copy-cell" onclick="copyCell(this)" title="${esc(codeStr)}">${esc(codeStr.length > 60 ? codeStr.slice(0,60)+'…' : codeStr)}</td>
              <td class="px-2 py-1 mono text-gray-500">${esc(gtin)}</td>
              <td class="px-2 py-1" style="color:${stConf.color}">${stConf.label}</td>
            </tr>`;
          }).join('')}
        </table>
      </div>
    </div>` : `<p class="text-xs text-gray-600">Коды маркировки не найдены в ответе API.<br>
      <span class="text-gray-700 mono">Ключи ответа: ${esc(Object.keys(d).join(', '))}</span></p>`;

  expRow.innerHTML = `<td colspan="6" class="px-4 py-3 border-t border-gray-900/50">
    <div class="grid grid-cols-2 md:grid-cols-4 gap-x-6 gap-y-1 text-xs mb-3">
      ${fields.map(([k,v]) => `
        <div class="text-gray-600 py-0.5">${esc(k)}</div>
        <div class="mono text-gray-300 py-0.5 truncate copy-cell" onclick="copyCell(this)" title="${esc(String(v))}">${esc(String(v))}</div>
      `).join('')}
    </div>
    ${posHtml}
    ${codesHtml}
  </td>`;
}

function showRcpErr(msg) {
  document.getElementById('rcp-err-text').textContent = msg;
  document.getElementById('rcp-err').classList.remove('hidden');
}

async function exportReceiptsExcel() {
  if (!rcpAllData.length) return;
  try {
    const res = await fetch('/api/receipts-excel', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({ receipts: rcpAllData })
    });
    if (!res.ok) { alert('Ошибка: ' + res.statusText); return; }
    const blob = await res.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url; a.download = 'cheki-kkt.xlsx'; a.click();
    URL.revokeObjectURL(url);
  } catch(e) { alert('Ошибка: ' + e.message); }
}

async function downloadAllRcpCodes() {
  if (!rcpAllData.length) return;
  const progWrap = document.getElementById('rcp-bulk-progress-wrap');
  const progText = document.getElementById('rcp-bulk-progress-text');
  const btn = document.getElementById('btn-rcp-bulk');
  btn.disabled = true;
  progWrap.classList.remove('hidden');

  const receiptsWithCodes = [];
  const total = rcpAllData.length;

  for (let i = 0; i < total; i++) {
    const r = rcpAllData[i];
    const rid = r.id || r.receiptId || r.checkId || '';
    progText.textContent = `Загружаю коды: чек ${i+1} из ${total}${rid ? ' ('+rid+')' : ''}`;

    if (!rid) continue;
    try {
      const res = await fetch('/api/rcp-detail/' + encodeURIComponent(rid));
      if (!res.ok) continue;
      const detail = await res.json();
      const d = detail.receipt || detail;

      let codes = [];
      for (const key of ['markingCodes', 'codes', 'cises', 'markCodes', 'items']) {
        if (Array.isArray(d[key]) && d[key].length > 0) { codes = d[key]; break; }
      }

      receiptsWithCodes.push({
        receiptId:  rid,
        externalId: r.externalId || d.externalId || '',
        type:       r.type || d.type || '',
        codes:      codes.map(c => ({
          code:  typeof c === 'string' ? c : (c.code || c.cis || c.markCode || c.identificationCode || JSON.stringify(c)),
          gtin:  typeof c === 'object' ? (c.gtin || c.productGtin || '') : '',
          state: typeof c === 'object' ? (c.state || c.status || c.markingCodeState || '') : '',
        })),
      });
    } catch(e) { /* skip failed receipts */ }
  }

  progText.textContent = 'Формирую Excel...';
  try {
    const res = await fetch('/api/rcp-bulk-codes-excel', {
      method: 'POST',
      headers: {'Content-Type':'application/json'},
      body: JSON.stringify({ receipts: receiptsWithCodes }),
    });
    if (!res.ok) { alert('Ошибка формирования Excel: ' + res.statusText); return; }
    const blob = await res.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url; a.download = 'cheki-kkt-vse-kody.xlsx'; a.click();
    URL.revokeObjectURL(url);
  } catch(e) { alert('Ошибка: ' + e.message); }
  finally {
    progWrap.classList.add('hidden');
    btn.disabled = rcpAllData.length === 0;
  }
}

// ════════════════════════════════════════════════════════════════════════════
// ── EXPORT TAB ────────────────────────────────────────────────────────────
// ════════════════════════════════════════════════════════════════════════════

async function startExport() {
  if (!S.token) { showExpErr('Введите токен.'); return; }
  document.getElementById('exp-err').classList.add('hidden');

  const gtin = document.getElementById('exp-gtin').value.trim();
  if (!gtin) { showExpErr('GTIN обязателен — API не поддерживает выгрузку без указания GTIN.'); return; }
  const body = { token: S.token, gtin };
  if (S.expStatusFilter.size) body.status = [...S.expStatusFilter];
  if (S.expPkgFilter.size)    body.packageType = [...S.expPkgFilter];

  const emFrom = document.getElementById('exp-em-from').value;
  const emTo   = document.getElementById('exp-em-to').value;
  const exFrom = document.getElementById('exp-ex-from').value;
  const exTo   = document.getElementById('exp-ex-to').value;
  if (emFrom) body.emissionDateFrom   = emFrom + 'T00:00:00Z';
  if (emTo)   body.emissionDateTo     = emTo   + 'T23:59:59Z';
  if (exFrom) body.expirationDateFrom = exFrom + 'T00:00:00Z';
  if (exTo)   body.expirationDateTo   = exTo   + 'T23:59:59Z';

  try {
    const res = await fetch('/api/export', {
      method: 'POST',
      headers: {'Content-Type':'application/json'},
      body: JSON.stringify(body)
    });
    const data = await res.json();
    if (!res.ok) { showExpErr(data.error || res.statusText); return; }
    S.exportJobId = data.id;
    document.getElementById('exp-job-id').textContent = data.id;
    document.getElementById('exp-job-wrap').classList.remove('hidden');
    document.getElementById('exp-download-btn').classList.add('hidden');
    setExpStep('CREATED');
    pollExportStatus();
  } catch(e) {
    showExpErr('Сетевая ошибка: ' + e.message);
  }
}

function setExpStep(status) {
  const sc = STATUS_DOC[status] || { label:status, bg:'#111', color:'#6b7280', border:'#1f2937' };
  document.getElementById('exp-job-badge').textContent = sc.label;
  document.getElementById('exp-job-badge').style.cssText = `background:${sc.bg};color:${sc.color};border-color:${sc.border}`;

  ['pending','processing','done'].forEach(s => {
    document.getElementById('step-'+s).style.cssText = '';
    document.getElementById('step-'+s).className = 'px-2 py-1 rounded border border-gray-800 text-gray-600';
  });
  if (status === 'CREATED')      { document.getElementById('step-pending').style.cssText    = 'color:#fbbf24;border-color:#92400e'; }
  if (status === 'IN_PROCESSING'){ document.getElementById('step-processing').style.cssText = 'color:#60a5fa;border-color:#1e3a5f'; }
  if (status === 'SUCCESS')      { document.getElementById('step-done').style.cssText        = 'color:#22c55e;border-color:#14532d'; }
  if (status === 'ERROR' || status === 'EXPIRED') { document.getElementById('step-done').style.cssText = 'color:#f87171;border-color:#450a0a'; }

  document.getElementById('exp-job-hint').textContent =
    status === 'SUCCESS'  ? '' :
    status === 'ERROR'    ? 'Ошибка при формировании выгрузки.' :
    status === 'EXPIRED'  ? 'Задание просрочено — результат удалён.' :
    'Задание выполняется на сервере. Можно ждать здесь.';
}

async function _doExportPoll() {
  if (!S.exportJobId) return;
  try {
    const res = await fetch(`/api/export/${S.exportJobId}/status?token=${encodeURIComponent(S.token)}`);
    const data = await res.json();
    const status = typeof data === 'string' ? data : (data.status || String(data));
    S.exportPollCount = (S.exportPollCount || 0) + 1;
    document.getElementById('exp-poll-count').textContent = S.exportPollCount;
    document.getElementById('exp-raw-status').textContent = status;
    setExpStep(status);
    if (status === 'SUCCESS') {
      _stopExportPoll();
      document.getElementById('exp-download-btn').classList.remove('hidden');
      document.getElementById('exp-refresh-btn').classList.add('hidden');
    }
    if (status === 'ERROR') {
      _stopExportPoll();
      showExpErr('Задание завершилось с ошибкой.');
    }
    if (status === 'EXPIRED') {
      _stopExportPoll();
      showExpErr('Задание просрочено — файл результата был удалён с сервера.');
    }
  } catch(e) {
    document.getElementById('exp-raw-status').textContent = 'сетевая ошибка: ' + e.message;
  }
}

function _stopExportPoll() {
  if (S.exportPollTimer)  { clearInterval(S.exportPollTimer);  S.exportPollTimer  = null; }
  if (S.exportElapsedTimer) { clearInterval(S.exportElapsedTimer); S.exportElapsedTimer = null; }
}

function pollExportStatus() {
  _stopExportPoll();
  S.exportPollCount = 0;
  const startTime = Date.now();
  S.exportElapsedTimer = setInterval(() => {
    const sec = Math.floor((Date.now() - startTime) / 1000);
    const el = document.getElementById('exp-elapsed');
    if (el) el.textContent = sec < 60 ? `${sec} сек` : `${Math.floor(sec/60)} мин ${sec%60} сек`;
  }, 1000);
  _doExportPoll();
  S.exportPollTimer = setInterval(_doExportPoll, 3000);
}

function checkExportNow() { _doExportPoll(); }

function downloadExport() {
  if (!S.exportJobId) return;
  window.location.href = `/api/export/${S.exportJobId}/result?token=${encodeURIComponent(S.token)}`;
}

function showExpErr(msg) {
  document.getElementById('exp-err-text').textContent = msg;
  document.getElementById('exp-err').classList.remove('hidden');
}

// ════════════════════════════════════════════════════════════════════════════
// ── RECONCILE TAB ────────────────────────────────────────────────────────
// ════════════════════════════════════════════════════════════════════════════

// Groups: ok / status / owner / missing
const REC_GROUP = {
  ok:      { label:'В обороте',    icon:'✅', color:'#22c55e' },
  status:  { label:'Статус не тот',icon:'⚠️', color:'#fbbf24' },
  owner:   { label:'Не ваш код',   icon:'❌', color:'#f87171' },
  missing: { label:'Не найден',    icon:'❓', color:'#6b7280' },
};

function classifyItem(item, myInn) {
  if (!item.status) return 'missing';
  if (myInn && item.issuerShortInfo?.issuerTin && item.issuerShortInfo.issuerTin !== myInn) return 'owner';
  if (item.status !== 'INTRODUCED') return 'status';
  return 'ok';
}

function onRecInput() {
  const codes = parseCodes(document.getElementById('rec-input').value);
  document.getElementById('rec-count').textContent = `${codes.length} кодов`;
}

function loadRecFile(input) {
  const file = input.files[0];
  if (!file) return;
  const reader = new FileReader();
  reader.onload = e => {
    const codes = e.target.result.split(/\r?\n/).map(s=>s.trim()).filter(s=>s.length>0);
    const unique = [...new Set(codes)];
    document.getElementById('rec-input').value = unique.join('\n');
    document.getElementById('rec-count').textContent = `${unique.length} кодов`;
    input.value = '';
  };
  reader.readAsText(file,'UTF-8');
}

async function runReconcile() {
  if (!S.token) { showRecErr('Введите токен.'); return; }
  const raw = document.getElementById('rec-input').value.trim();
  if (!raw) return;

  const codes = [...new Set(parseCodes(raw))];
  if (!codes.length) { showRecErr('Нет кодов для проверки.'); return; }
  const myInn = document.getElementById('rec-inn').value.trim();

  S.reconcileResults = [];
  S.reconcileFilter = 'all';
  document.getElementById('rec-tbody').innerHTML = '';
  document.getElementById('rec-err').classList.add('hidden');
  document.getElementById('rec-summary').classList.add('hidden');
  document.getElementById('rec-filter-wrap').classList.add('hidden');
  document.getElementById('rec-result-wrap').classList.add('hidden');
  document.getElementById('rec-progress-wrap').classList.remove('hidden');
  document.getElementById('btn-rec').disabled = true;
  setRecFilterPill('all');

  const chunks = [];
  for (let i=0; i<codes.length; i+=1000) chunks.push(codes.slice(i,i+1000));
  let done = 0;
  updateRecProgress(0, chunks.length);

  for (const chunk of chunks) {
    try {
      const res = await fetch('/api/info', {
        method:'POST', headers:{'Content-Type':'application/json'},
        body: JSON.stringify({ codes:chunk, token:S.token })
      });
      if (!res.ok) { const e=await res.json().catch(()=>{}); showRecErr(`Ошибка API: ${e?.error||res.statusText}`); break; }
      const data = await res.json();
      // Classify each result
      data.forEach(item => {
        const group = classifyItem(item, myInn);
        S.reconcileResults.push({ ...item, _group: group });
      });
      appendRecRows(data, myInn);
    } catch(e) { showRecErr('Сетевая ошибка: ' + e.message); break; }
    done++;
    updateRecProgress(done, chunks.length);
    if (done < chunks.length) await delay(S.chunkDelay);
  }

  document.getElementById('rec-progress-wrap').classList.add('hidden');
  document.getElementById('btn-rec').disabled = false;
  document.getElementById('btn-rec-csv').disabled = false;
  document.getElementById('btn-rec-json').disabled = false;
  document.getElementById('rec-summary').classList.remove('hidden');
  document.getElementById('rec-filter-wrap').classList.remove('hidden');
  document.getElementById('rec-result-wrap').classList.remove('hidden');
  updateRecSummary();
  updateRecLabel();
}

function appendRecRows(items, myInn) {
  const tbody = document.getElementById('rec-tbody');
  items.forEach(item => {
    const group = classifyItem(item, myInn);
    const g = REC_GROUP[group];
    const sc = STATUS[item.status] || { color:'#6b7280' };
    const exp = item.expirationDate ? new Date(item.expirationDate).toLocaleDateString('ru-RU') : '—';
    const expired = item.expirationDate && new Date(item.expirationDate) < new Date();
    const tin = item.issuerShortInfo?.issuerTin || '—';

    const tr = document.createElement('tr');
    tr.className = 'result-row row-enter';
    tr.dataset.group = group;
    tr.innerHTML = `
      <td class="px-3 py-2.5 mono text-xs text-gray-300 max-w-[220px] truncate copy-cell"
          onclick="copyCell(this)" title="${esc(item.code)}">${esc(item.code||'—')}</td>
      <td class="px-3 py-2.5">
        ${item.status
          ? `<span class="status-badge" style="background:${sc.bg||'#111'};color:${sc.color};border-color:${sc.border||'#1f2937'}">${sc.label||item.status}</span>`
          : '<span class="text-gray-700 text-xs">—</span>'}
      </td>
      <td class="px-3 py-2.5 mono text-xs text-gray-400 copy-cell" onclick="copyCell(this)">${esc(tin)}</td>
      <td class="px-3 py-2.5 text-sm" style="color:${g.color}">${g.icon} ${g.label}</td>
      <td class="px-3 py-2.5 text-xs ${expired?'text-red-400':'text-gray-500'}">${exp}</td>`;
    tbody.appendChild(tr);
  });
}

function setRecFilter(f) {
  S.reconcileFilter = f;
  setRecFilterPill(f);
  document.querySelectorAll('#rec-tbody tr[data-group]').forEach(tr => {
    tr.classList.toggle('hidden', f!=='all' && tr.dataset.group!==f);
  });
  updateRecLabel();
}
function setRecFilterPill(f) {
  document.querySelectorAll('#rec-filter-wrap .filter-pill').forEach(p =>
    p.classList.toggle('active', p.dataset.filter===f));
}

function updateRecSummary() {
  const counts = { ok:0, status:0, owner:0, missing:0 };
  S.reconcileResults.forEach(r => counts[r._group]++);
  document.getElementById('rs-total').textContent = S.reconcileResults.length;
  document.getElementById('rs-ok').textContent    = counts.ok;
  document.getElementById('rs-warn').textContent  = counts.status + counts.owner;
  document.getElementById('rs-miss').textContent  = counts.missing;
}

function updateRecLabel() {
  const vis = document.querySelectorAll('#rec-tbody tr[data-group]:not(.hidden)').length;
  document.getElementById('rec-result-label').textContent = `Показано: ${vis} из ${S.reconcileResults.length}`;
}

function updateRecProgress(done, total) {
  const pct = total ? Math.round(done/total*100) : 0;
  document.getElementById('rec-progress-bar').style.width = `${pct}%`;
  const rem = total - done;
  const etaSec = Math.ceil(rem * S.chunkDelay / 1000);
  const eta = done < total && etaSec > 1 ? ` · ~${etaSec} сек` : '';
  document.getElementById('rec-progress-text').textContent = `${done} / ${total} чанков${eta}`;
}

function showRecErr(msg) {
  document.getElementById('rec-err-text').textContent = msg;
  document.getElementById('rec-err').classList.remove('hidden');
}

function resetReconcile() {
  S.reconcileResults = []; S.reconcileFilter = 'all';
  document.getElementById('rec-input').value = '';
  document.getElementById('rec-count').textContent = '0 кодов';
  document.getElementById('rec-tbody').innerHTML = '';
  ['rec-summary','rec-filter-wrap','rec-result-wrap','rec-progress-wrap','rec-err'].forEach(id =>
    document.getElementById(id).classList.add('hidden'));
  document.getElementById('btn-rec-csv').disabled = true;
  document.getElementById('btn-rec-json').disabled = true;
}

function exportRecCSV() {
  const bad = S.reconcileResults.filter(r => r._group !== 'ok');
  if (!bad.length) return;
  const hdr = ['code','group','status','issuerTin','expirationDate'];
  const rows = [hdr.join(',')];
  bad.forEach(r => rows.push([
    r.code, r._group, r.status||'', r.issuerShortInfo?.issuerTin||'', r.expirationDate||''
  ].join(',')));
  download('﻿'+rows.join('\r\n'), 'reconcile-problems.csv', 'text/csv;charset=utf-8;');
}

function exportRecJSON() {
  download(JSON.stringify(S.reconcileResults,null,2), 'reconcile-all.json', 'application/json');
}

// ── Helpers ────────────────────────────────────────────────────────────────
function esc(s) {
  return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}
</script>
</body>
</html>"""

if __name__ == "__main__":
    app.run(debug=True, port=8000)
